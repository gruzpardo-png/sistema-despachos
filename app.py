
import os
import shutil
import json
import sqlite3
import base64
import re
import mimetypes
import unicodedata
from io import BytesIO
from datetime import datetime, date
from zoneinfo import ZoneInfo
from functools import wraps
from difflib import SequenceMatcher

from flask import (
    Flask, request, redirect, url_for, session, flash,
    send_file, render_template_string
)
from markupsafe import Markup
from werkzeug.security import generate_password_hash, check_password_hash
from werkzeug.utils import secure_filename
from openpyxl import Workbook


APP_NAME = "Ferretería Cloud Tool"
APP_VERSION = "v4.6 Chat Ventas Integrado"
DB_PATH = os.environ.get("DATABASE_PATH", "ferreteria_cloud_tool.db")
SECRET_KEY = os.environ.get("SECRET_KEY", "cambiar-esta-clave-en-render")


app = Flask(__name__)
app.secret_key = SECRET_KEY
app.config['MAX_CONTENT_LENGTH'] = 25 * 1024 * 1024


# ============================================================
# UTILIDADES BASE
# ============================================================

CHILE_TZ = ZoneInfo("America/Santiago")


def chile_now():
    return datetime.now(CHILE_TZ)


def now_str():
    return chile_now().strftime("%Y-%m-%d %H:%M:%S")


def today_str():
    return chile_now().date().isoformat()


def database_directory():
    folder = os.path.dirname(os.path.abspath(DB_PATH))
    return folder if folder else "."


def backup_database_if_exists(label="manual"):
    """
    Crea un respaldo físico de la base SQLite si existe.
    No borra ni modifica la base principal.
    Mantiene los últimos 20 respaldos para no llenar el disco.
    """
    try:
        if not os.path.exists(DB_PATH):
            return None
        if os.path.getsize(DB_PATH) == 0:
            return None

        backup_dir = os.path.join(database_directory(), "backups")
        os.makedirs(backup_dir, exist_ok=True)

        stamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        filename = f"ferreteria_cloud_tool_{label}_{stamp}.db"
        destination = os.path.join(backup_dir, filename)

        shutil.copy2(DB_PATH, destination)

        backups = sorted(
            [
                os.path.join(backup_dir, f)
                for f in os.listdir(backup_dir)
                if f.endswith(".db")
            ],
            key=os.path.getmtime,
            reverse=True
        )

        for old_backup in backups[20:]:
            try:
                os.remove(old_backup)
            except Exception:
                pass

        return destination
    except Exception:
        return None


def ensure_database_parent():
    parent = os.path.dirname(os.path.abspath(DB_PATH))
    if parent and not os.path.exists(parent):
        try:
            os.makedirs(parent, exist_ok=True)
            print(f"[DB] Carpeta creada para SQLite: {parent}", flush=True)
        except Exception as exc:
            print("[DB ERROR] No se pudo crear/abrir la carpeta de la base de datos.", flush=True)
            print(f"[DB ERROR] DATABASE_PATH={DB_PATH}", flush=True)
            print(f"[DB ERROR] Carpeta esperada={parent}", flush=True)
            print("[DB ERROR] En Render revisa: Disks -> Mount Path y Environment -> DATABASE_PATH.", flush=True)
            print(f"[DB ERROR] Detalle original: {exc}", flush=True)
            raise


def db():
    ensure_database_parent()
    try:
        conn = sqlite3.connect(DB_PATH, timeout=30)
    except sqlite3.OperationalError as exc:
        print("[DB ERROR] SQLite no pudo abrir la base de datos.", flush=True)
        print(f"[DB ERROR] DATABASE_PATH={DB_PATH}", flush=True)
        print("[DB ERROR] Causa probable: no existe Disk persistente, mount path incorrecto o permiso insuficiente.", flush=True)
        print("[DB ERROR] Solución recomendada: Disk Mount Path /data y DATABASE_PATH=/data/ferreteria_cloud_tool.db", flush=True)
        print(f"[DB ERROR] Detalle original: {exc}", flush=True)
        raise
    conn.row_factory = sqlite3.Row
    return conn


def execute(sql, params=()):
    with db() as conn:
        conn.execute(sql, params)
        conn.commit()


def insert_and_get_id(sql, params=()):
    with db() as conn:
        cur = conn.execute(sql, params)
        conn.commit()
        return cur.lastrowid


def query_all(sql, params=()):
    with db() as conn:
        return conn.execute(sql, params).fetchall()


def query_one(sql, params=()):
    with db() as conn:
        return conn.execute(sql, params).fetchone()


def money_to_float(value):
    if value is None:
        return 0.0
    text = str(value).replace("$", "").replace(" ", "").strip()
    if "," in text and "." in text:
        text = text.replace(".", "").replace(",", ".")
    elif "," in text:
        text = text.replace(",", ".")
    try:
        return float(text)
    except Exception:
        return 0.0


@app.template_filter("money")
def money(value):
    try:
        number = float(value or 0)
    except Exception:
        number = 0
    return "$" + f"{number:,.0f}".replace(",", ".")


@app.template_filter("percent")
def percent(value):
    try:
        number = float(value or 0)
    except Exception:
        number = 0
    return f"{number * 100:.1f}%".replace(".", ",")


@app.template_filter("date_short")
def date_short(value):
    if not value:
        return ""
    return str(value)[:10]


def table_columns(table_name):
    try:
        rows = query_all(f"PRAGMA table_info({table_name})")
        return {r["name"] for r in rows}
    except Exception:
        return set()


def add_column_if_missing(table_name, column_definition):
    column_name = column_definition.split()[0]
    if column_name not in table_columns(table_name):
        execute(f"ALTER TABLE {table_name} ADD COLUMN {column_definition}")


def get_config_list(clave, default=None):
    default = default or []
    row = query_one("SELECT valor FROM system_config WHERE clave = ?", (clave,))
    if not row:
        return default
    try:
        data = json.loads(row["valor"])
        return data if isinstance(data, list) else default
    except Exception:
        return default


def set_config_list(clave, values):
    clean = []
    for v in values:
        v = str(v).strip()
        if v and v not in clean:
            clean.append(v)
    execute("""
        INSERT INTO system_config (clave, valor, updated_at)
        VALUES (?, ?, ?)
        ON CONFLICT(clave) DO UPDATE SET valor = excluded.valor, updated_at = excluded.updated_at
    """, (clave, json.dumps(clean, ensure_ascii=False), now_str()))


# ============================================================
# UTILIDADES PRODUCTOS / VENTAS / IA
# ============================================================

def iva_rate():
    raw = os.environ.get("IVA_RATE", "0.19")
    try:
        return float(str(raw).replace(",", "."))
    except Exception:
        return 0.19


def openai_model_name():
    # Configurable desde Render. Usa el modelo exacto que tenga disponible tu cuenta.
    return os.environ.get("OPENAI_MODEL", "gpt-5.4-mini")


def openai_is_configured():
    return bool(os.environ.get("OPENAI_API_KEY"))


def normalize_text(value):
    value = "" if value is None else str(value)
    value = unicodedata.normalize("NFKD", value)
    value = "".join(c for c in value if not unicodedata.combining(c))
    value = value.upper()
    value = re.sub(r"[^A-Z0-9]+", " ", value)
    return re.sub(r"\s+", " ", value).strip()


def parse_number(value, default=0.0):
    if value is None:
        return default
    if isinstance(value, (int, float)):
        return float(value)
    text = str(value).strip()
    if not text:
        return default
    text = text.replace("$", "").replace(" ", "").replace("\xa0", "")
    # Formato chileno usual: 1.989,68
    if "," in text:
        text = text.replace(".", "").replace(",", ".")
    else:
        # Si viene 1,989 desde un sistema anglo, elimina separador miles.
        if text.count(".") > 1:
            parts = text.split(".")
            text = "".join(parts[:-1]) + "." + parts[-1]
    try:
        return float(text)
    except Exception:
        return default


def parse_bool_si(value):
    text = normalize_text(value)
    return 1 if text in {"SI", "S", "YES", "TRUE", "1", "ACTIVO", "ACTIVA"} else 0


def parse_quantity(value, default=1.0):
    qty = parse_number(value, default)
    if qty <= 0:
        return default
    return qty


def venta_neta_desde_bruto(precio_bruto):
    return float(precio_bruto or 0) / (1 + iva_rate())


def contribucion_unitaria(precio_compra_neto, precio_venta_bruto):
    return venta_neta_desde_bruto(precio_venta_bruto) - float(precio_compra_neto or 0)


def margen_pct(precio_compra_neto, precio_venta_bruto):
    venta_neta = venta_neta_desde_bruto(precio_venta_bruto)
    if venta_neta <= 0:
        return 0
    return contribucion_unitaria(precio_compra_neto, precio_venta_bruto) / venta_neta



SALES_STOPWORDS = {
    "DE", "DEL", "LA", "EL", "LOS", "LAS", "PARA", "POR", "CON", "UN", "UNA",
    "UNIDAD", "UNIDADES", "UND", "UD", "U", "X", "NECESITO", "QUIERO", "DAME",
    "COTIZAR", "COTIZACION", "COTIZACIÓN", "LISTA", "PRODUCTO", "PRODUCTOS",
    "HOLA", "BUENAS", "BUENOS", "DIAS", "DIA", "TARDES", "NOCHES", "FAVOR",
    "PORFA", "PORFAVOR", "QUISEIRA", "QUISIERA", "SGTE", "SIGUIENTE", "SIGTE",
    "LO", "ESTO", "ESTOS", "LAS", "LOS", "MM", "CM", "MT", "MTS", "METRO",
    "METROS", "BTO", "BRUTO", "NETO"
}


def producto_por_codigo(codigo):
    if not codigo:
        return None
    return query_one("""
        SELECT * FROM productos
        WHERE codigo_producto = ?
        LIMIT 1
    """, (str(codigo).strip(),))


def canonical_sales_line(text):
    text = str(text or "").strip(" -•\t")
    text = re.sub(r"\s+", " ", text).strip()
    return text


def is_noise_line(line):
    n = normalize_text(line)
    if not n:
        return True
    noise_exact = {
        "HOLA", "BUENAS", "BUENOS DIAS", "BUENAS TARDES", "BUENAS NOCHES",
        "QUISIERA COTIZAR", "QUIERO COTIZAR", "LO SGTE", "LO SIGUIENTE",
        "LO SGTE", "GRACIAS", "SALUDOS"
    }
    if n in noise_exact:
        return True
    if len(n.split()) <= 2 and any(w in n for w in ["HOLA", "COTIZAR", "SGTE", "SIGUIENTE"]):
        return True
    return False


def is_dimension_or_spec_line(line):
    n = normalize_text(line)
    if not n:
        return False
    # líneas como "5 mm", "0,4 mm x 3 mt", "0,7 mm x 3,66 mt", "2x3 bto"
    if re.fullmatch(r"[0-9]+(?: [0-9]+)?(?: X [0-9]+)?(?: MM|CM|MT|MTS|M)?(?: X [0-9]+(?: MM|CM|MT|MTS|M)?)?(?: BTO)?", n):
        return True
    if re.fullmatch(r"[0-9]+X[0-9]+(?: BTO)?", n.replace(" ", "")):
        return True
    if any(unit in n.split() for unit in ["MM", "CM", "MT", "MTS"]) and len(n.split()) <= 6:
        return True
    return False


def normalize_dimension_tokens(text):
    text = str(text or "").upper()
    text = text.replace(",", ".")
    text = re.sub(r"(\d+)\s*[Xx]\s*(\d+)", r"\1X\2", text)
    text = re.sub(r"(\d+(?:\.\d+)?)\s*(MM|CM|MT|MTS|M)\b", r"\1 \2", text)
    return text


def tokenize_search(text):
    norm = normalize_dimension_tokens(normalize_text(text))
    raw_tokens = norm.split()
    tokens = []
    for t in raw_tokens:
        if t in SALES_STOPWORDS:
            continue
        if re.fullmatch(r"\d+", t):
            # número solo no sirve para buscar; se usa con contexto pero no como token primario.
            continue
        if len(t) < 3 and not re.fullmatch(r"\d+X\d+", t):
            continue
        tokens.append(t)
    clean = []
    for t in tokens:
        if t not in clean:
            clean.append(t)
    return clean[:10]


def token_similarity(a, b):
    if not a or not b:
        return 0
    if a == b:
        return 1
    if len(a) >= 5 and len(b) >= 5:
        return SequenceMatcher(None, a, b).ratio()
    return 0


def producto_score(query_text, producto):
    query_norm = normalize_text(query_text)
    desc_norm = producto["descripcion_busqueda"] or normalize_text(producto["descripcion"])
    q_tokens = tokenize_search(query_text)
    d_tokens = desc_norm.split()

    if not q_tokens:
        return 0, {"overlap": 0, "tokens": []}

    exact_phrase = query_norm and len(query_norm) >= 8 and query_norm in desc_norm
    overlap = 0
    fuzzy_overlap = 0
    matched = []

    for qt in q_tokens:
        matched_exact = qt in d_tokens or (len(qt) >= 5 and qt in desc_norm)
        matched_fuzzy = False
        if not matched_exact and len(qt) >= 5:
            for dt in d_tokens:
                if token_similarity(qt, dt) >= 0.86:
                    matched_fuzzy = True
                    break
        if matched_exact:
            overlap += 1
            matched.append(qt)
        elif matched_fuzzy:
            fuzzy_overlap += 1
            matched.append(qt + "~")

    total_overlap = overlap + fuzzy_overlap
    ratio = total_overlap / max(len(q_tokens), 1)

    score = 0
    if exact_phrase:
        score += 55

    # base por cobertura de términos reales
    score += ratio * 65
    score += min(total_overlap, 4) * 8
    score += fuzzy_overlap * 4

    # penaliza si el producto solo calza por medidas/códigos sueltos pero no por material/familia
    strong_query = [t for t in q_tokens if not re.search(r"\d", t) and len(t) >= 4]
    strong_match = 0
    for t in strong_query:
        if t in desc_norm or any(token_similarity(t, dt) >= 0.86 for dt in d_tokens):
            strong_match += 1

    if strong_query and strong_match == 0:
        score -= 45
    if len(q_tokens) >= 2 and total_overlap < 2 and not exact_phrase:
        score -= 30

    # pequeño ajuste por activo/stock, nunca suficiente para forzar un mal match
    try:
        if int(producto["activo"] or 0) == 1:
            score += 3
        if float(producto["stock"] or 0) > 0:
            score += 2
    except Exception:
        pass

    return round(max(score, 0), 1), {"overlap": total_overlap, "tokens": matched}


def buscar_producto_local(descripcion="", codigo_producto=None):
    codigo = str(codigo_producto or "").strip()
    if codigo:
        exact = producto_por_codigo(codigo)
        if exact:
            return exact, 100

        # Solo permite búsqueda parcial por código si el código es suficientemente específico.
        if len(codigo) >= 6:
            like = query_one("""
                SELECT * FROM productos
                WHERE codigo_producto LIKE ?
                ORDER BY activo DESC, stock DESC
                LIMIT 1
            """, (f"%{codigo}%",))
            if like:
                return like, 82

    tokens = tokenize_search(descripcion)
    if not tokens:
        return None, 0

    # Buscar candidatos por OR amplio pero luego aplicar score estricto.
    clauses = []
    params = []
    for t in tokens[:6]:
        clauses.append("descripcion_busqueda LIKE ?")
        params.append(f"%{t}%")

    rows = []
    if clauses:
        rows = query_all(f"""
            SELECT * FROM productos
            WHERE activo = 1 AND ({' OR '.join(clauses)})
            LIMIT 250
        """, params)

    # Si hay typos, usa términos de categoría genéricos para acotar.
    if not rows:
        broad_tokens = [t for t in tokens if not re.search(r"\d", t)][:3]
        if broad_tokens:
            clauses = ["descripcion_busqueda LIKE ?" for _ in broad_tokens]
            params = [f"%{t[:5]}%" for t in broad_tokens]
            rows = query_all(f"""
                SELECT * FROM productos
                WHERE activo = 1 AND ({' OR '.join(clauses)})
                LIMIT 250
            """, params)

    if not rows:
        return None, 0

    best = None
    best_score = -1
    for row in rows:
        score, _meta = producto_score(descripcion, row)
        if score > best_score:
            best = row
            best_score = score

    # Umbral alto para no vender productos equivocados.
    # 72+ se considera match confiable.
    if best and best_score >= 72:
        return best, best_score

    # 58-71 se devuelve como candidato a revisión, no como match confirmado.
    if best and best_score >= 58:
        return {"__candidato__": best, "__score__": best_score}, best_score

    return None, best_score


def extraer_items_local(texto):
    """
    Fallback sin IA. Convierte lista de materiales en líneas limpias.
    Une líneas de especificación como "5 mm" o "0,4 mm x 3 mt" con el producto anterior.
    """
    raw = texto or ""
    prelim = []
    for chunk in re.split(r"[\n;]+", raw):
        line = canonical_sales_line(chunk)
        if line:
            prelim.append(line)

    lines = []
    for line in prelim:
        if is_noise_line(line):
            continue
        if is_dimension_or_spec_line(line) and lines:
            lines[-1] = canonical_sales_line(lines[-1] + " " + line)
        else:
            lines.append(line)

    items = []
    for line in lines:
        qty = 1.0
        product_text = line

        # Cantidades explícitas al inicio: "3 planchas osb", "2 x cemento".
        m = re.search(r"^\s*(\d+(?:[,.]\d+)?)\s*(?:X|x|UN|UND|UDS|UNIDADES|UNIDAD)?\s+(.+)$", line)
        if m:
            possible_qty = parse_quantity(m.group(1), 1)
            rest = m.group(2).strip()
            # Evita confundir medidas como "0,4 mm x 3 mt" con cantidades.
            if not re.match(r"^(MM|CM|MT|MTS|M)\b", normalize_text(rest)):
                qty = possible_qty
                product_text = rest
        else:
            m = re.search(r"(.+?)\s+(?:X|x)\s*(\d+(?:[,.]\d+)?)\s*$", line)
            if m and not re.search(r"\b(MM|CM|MT|MTS|M)\b", normalize_text(line)):
                product_text = m.group(1).strip()
                qty = parse_quantity(m.group(2), 1)

        product_text = canonical_sales_line(product_text)
        if product_text and not is_noise_line(product_text):
            items.append({
                "codigo_producto": None,
                "descripcion": product_text,
                "cantidad": qty,
                "observacion": "extracción local"
            })

    return {"items": items, "notas": "Extracción local sin IA"}


def imagen_a_data_url(file_storage):
    if not file_storage or not file_storage.filename:
        return None
    raw = file_storage.read()
    if not raw:
        return None
    if len(raw) > 12 * 1024 * 1024:
        raise ValueError("La imagen supera el tamaño máximo permitido de 12 MB.")
    mime = file_storage.mimetype or mimetypes.guess_type(file_storage.filename)[0] or "image/jpeg"
    encoded = base64.b64encode(raw).decode("utf-8")
    return f"data:{mime};base64,{encoded}"


def extraer_json_desde_respuesta(raw):
    raw = (raw or "").strip()
    if not raw:
        return None
    try:
        return json.loads(raw)
    except Exception:
        pass
    start = raw.find("{")
    end = raw.rfind("}")
    if start >= 0 and end > start:
        try:
            return json.loads(raw[start:end + 1])
        except Exception:
            return None
    return None


def extraer_items_con_elias(texto, imagen_file=None):
    """
    Extrae items para cotización, pero no cotiza. La cotización se genera solo cuando
    el vendedor presiona el botón Generar cotización.
    """
    data_url = None
    if imagen_file and imagen_file.filename:
        data_url = imagen_a_data_url(imagen_file)

    if not openai_is_configured():
        return extraer_items_local(texto), "local", "OPENAI_API_KEY no configurada"

    try:
        from openai import OpenAI
        client = OpenAI()

        system_prompt = (
            "Eres Elias, asistente de ventas de Ferretería San Pedro. "
            "Tu tarea es EXTRAER líneas de productos desde una conversación o imagen para preparar una cotización. "
            "No busques productos en una base, no inventes códigos, no inventes precios. "
            "Ignora saludos, frases como 'quiero cotizar', 'lo siguiente', 'gracias'. "
            "Une especificaciones que vienen en líneas separadas con el producto anterior: por ejemplo 'Plancha fibrocemento' + '5 mm'. "
            "IMPORTANTE: dimensiones como 0,4 mm, 0,7 mm, 3 mt, 3,66 mt NO son cantidades; son parte de la descripción. "
            "Si no hay cantidad explícita, usa cantidad 1. "
            "Devuelve SOLO JSON válido sin markdown con estructura: "
            "{\"items\":[{\"codigo_producto\":null,\"descripcion\":\"producto limpio con medida\",\"cantidad\":1,\"observacion\":\"\"}],\"notas\":\"\"}."
        )

        user_text = "Extrae productos y cantidades desde esta conversación/lista:\n\n" + (texto or "")

        user_content = [{"type": "input_text", "text": user_text}]
        if data_url:
            user_content.append({"type": "input_image", "image_url": data_url})

        response = client.responses.create(
            model=openai_model_name(),
            input=[
                {"role": "developer", "content": [{"type": "input_text", "text": system_prompt}]},
                {"role": "user", "content": user_content},
            ],
            max_output_tokens=2500,
        )

        raw = getattr(response, "output_text", None) or str(response)
        data = extraer_json_desde_respuesta(raw)
        if not data or "items" not in data:
            return extraer_items_local(texto), "local", f"No se pudo interpretar JSON de Elias. Respuesta parcial: {raw[:500]}"

        # Limpieza posterior: elimina líneas basura.
        clean_items = []
        for item in data.get("items", []):
            desc = canonical_sales_line(item.get("descripcion") or item.get("producto") or "")
            if desc and not is_noise_line(desc):
                clean_items.append({
                    "codigo_producto": item.get("codigo_producto") or item.get("codigo"),
                    "descripcion": desc,
                    "cantidad": parse_quantity(item.get("cantidad"), 1),
                    "observacion": item.get("observacion") or "",
                })
        data["items"] = clean_items
        return data, "openai", raw

    except Exception as exc:
        return extraer_items_local(texto), "local", f"Error IA: {exc}"


def respuesta_chat_elias(mensaje, contexto, imagen_file=None):
    """
    Respuesta conversacional. No genera cotización. Puede ayudar a aclarar y ordenar el pedido.
    """
    data_url = None
    if imagen_file and imagen_file.filename:
        data_url = imagen_a_data_url(imagen_file)

    if not openai_is_configured():
        data = extraer_items_local(mensaje)
        items = data.get("items", [])
        if items:
            lista = "\n".join([f"- {i['descripcion']} · cantidad {i['cantidad']}" for i in items])
            return "Detecté estos productos de forma local:\n" + lista + "\n\nCuando esté correcto, presiona el botón Generar cotización.", "local"
        return "Escríbeme la lista de productos o adjunta una imagen del pedido. Cuando esté clara, usa el botón Generar cotización.", "local"

    try:
        from openai import OpenAI
        client = OpenAI()
        system_prompt = (
            "Eres Elias, asistente de ventas interno de Ferretería San Pedro. "
            "Conversas con el vendedor para ordenar el pedido del cliente. "
            "NO generes cotización ni precios automáticamente. "
            "NO inventes códigos, precios ni stock. "
            "Si detectas productos, responde con formato de chat claro: 'Pedido ordenado:' y una lista por líneas. "
            "Une medidas separadas con el producto anterior. Ejemplo: 'Plancha fibrocemento' + '5 mm' debe quedar 'Plancha fibrocemento 5 mm'. "
            "Dimensiones como 0,4 mm x 3 mt NO son cantidades, son especificaciones. "
            "Si el pedido es ambiguo, pregunta una sola cosa concreta. "
            "Termina siempre con: 'Cuando esté correcto, presiona Generar cotización'."
        )

        user_content = [{"type": "input_text", "text": f"Contexto anterior:\n{contexto}\n\nNuevo mensaje:\n{mensaje}"}]
        if data_url:
            user_content.append({"type": "input_image", "image_url": data_url})

        response = client.responses.create(
            model=openai_model_name(),
            input=[
                {"role": "developer", "content": [{"type": "input_text", "text": system_prompt}]},
                {"role": "user", "content": user_content},
            ],
            max_output_tokens=1200,
        )
        return (getattr(response, "output_text", None) or str(response)).strip(), "openai"
    except Exception as exc:
        error_text = str(exc)
        if "No module named 'openai'" in error_text:
            return "No pude consultar IA porque falta instalar la librería OpenAI en Render. Revisa que requirements.txt tenga openai>=1.100.0 y redeploya con Clear build cache. Mientras tanto puedes generar cotización con extracción local.", "local"
        return f"No pude consultar IA en este momento ({exc}). Puedes pegar la lista y luego presionar Generar cotización para usar extracción local.", "local"


def obtener_sesion_ventas():
    sid = session.get("ventas_chat_sesion_id")
    if sid:
        row = query_one("SELECT * FROM ventas_chat_sesiones WHERE id = ?", (sid,))
        if row:
            return row

    user = current_user()
    sid = insert_and_get_id("""
        INSERT INTO ventas_chat_sesiones (titulo, cliente, telefono, estado, usuario_id, usuario_nombre, created_at, updated_at)
        VALUES (?, '', '', 'Abierta', ?, ?, ?, ?)
    """, (
        "Nueva conversación Elias",
        user["id"] if user else None,
        user["full_name"] if user else "Sistema",
        now_str(),
        now_str(),
    ))
    session["ventas_chat_sesion_id"] = sid
    return query_one("SELECT * FROM ventas_chat_sesiones WHERE id = ?", (sid,))


def mensajes_sesion_ventas(sesion_id):
    return query_all("SELECT * FROM ventas_chat_mensajes WHERE sesion_id = ? ORDER BY id", (sesion_id,))


def guardar_mensaje_ventas(sesion_id, rol, contenido, ai_raw="", tiene_imagen=0):
    mid = insert_and_get_id("""
        INSERT INTO ventas_chat_mensajes (sesion_id, rol, contenido, ai_raw, tiene_imagen, created_at)
        VALUES (?, ?, ?, ?, ?, ?)
    """, (sesion_id, rol, contenido or "", ai_raw or "", 1 if tiene_imagen else 0, now_str()))
    execute("UPDATE ventas_chat_sesiones SET updated_at=? WHERE id=?", (now_str(), sesion_id))
    return mid


def contexto_conversacion_ventas(sesion_id):
    mensajes = mensajes_sesion_ventas(sesion_id)
    partes = []
    for m in mensajes[-30:]:
        rol = "Vendedor" if m["rol"] == "user" else "Elias"
        partes.append(f"{rol}: {m['contenido']}")
    return "\n".join(partes)


def extraer_items_desde_sesion_ventas(sesion_id):
    contexto = contexto_conversacion_ventas(sesion_id)
    return extraer_items_con_elias(contexto, None)


def row_to_dict(row):
    return dict(row) if row else None


def cotizacion_chat_data(cotizacion_id):
    cot = query_one("SELECT * FROM cotizaciones WHERE id = ?", (cotizacion_id,))
    if not cot:
        return None
    items = query_all("SELECT * FROM cotizacion_items WHERE cotizacion_id = ? ORDER BY id", (cotizacion_id,))
    revisar_count = query_one(
        "SELECT COUNT(*) c FROM cotizacion_items WHERE cotizacion_id=? AND (encontrado=0 OR requiere_revision=1)",
        (cotizacion_id,)
    )["c"]
    return {
        "cot": row_to_dict(cot),
        "items": [row_to_dict(i) for i in items],
        "revisar_count": revisar_count,
    }


def mensajes_sesion_ventas_render(sesion_id):
    mensajes = mensajes_sesion_ventas(sesion_id)
    salida = []
    for m in mensajes:
        d = row_to_dict(m)
        d["quote"] = None
        if d.get("rol") == "quote":
            try:
                payload = json.loads(d.get("ai_raw") or "{}")
                d["quote"] = cotizacion_chat_data(payload.get("cotizacion_id"))
            except Exception:
                d["quote"] = None
        salida.append(d)
    return salida


def resumen_cotizacion_chat(cotizacion_id):
    data = cotizacion_chat_data(cotizacion_id)
    if not data:
        return "Cotización generada, pero no se pudo cargar el resumen."
    cot = data["cot"]
    revisar = data["revisar_count"]
    return (
        f"Cotización {cot['numero']} generada en el chat.\n"
        f"Total bruto confirmado: {money(cot['subtotal_bruto'])}\n"
        f"Venta neta: {money(cot['venta_neta_total'])}\n"
        f"Contribución: {money(cot['contribucion_total'])}\n"
        f"Margen: {percent(cot['margen_total_pct'])}\n"
        f"Revisión pendiente: {revisar} línea(s)."
    )


def format_clp(value):
    return money(value)


def generar_numero_cotizacion(cotizacion_id):
    return f"COT-{today_str().replace('-', '')}-{int(cotizacion_id):06d}"



def crear_cotizacion_desde_items(cliente, telefono, texto_original, items_extraidos, fuente, ai_raw, chat_session_id=None):
    user = current_user()
    now = now_str()

    cot_id = insert_and_get_id("""
        INSERT INTO cotizaciones (
            numero, cliente, telefono, texto_original, fuente, ai_raw,
            subtotal_bruto, venta_neta_total, costo_neto_total, contribucion_total, margen_total_pct,
            estado, usuario_id, usuario_nombre, created_at, updated_at
        )
        VALUES (?, ?, ?, ?, ?, ?, 0, 0, 0, 0, 0, 'Borrador generado', ?, ?, ?, ?)
    """, (
        "PENDIENTE",
        cliente.strip() if cliente else "",
        telefono.strip() if telefono else "",
        texto_original or "",
        fuente,
        ai_raw or "",
        user["id"] if user else None,
        user["full_name"] if user else "Sistema",
        now,
        now,
    ))

    try:
        add_column_if_missing("cotizaciones", "chat_session_id INTEGER")
        execute("UPDATE cotizaciones SET chat_session_id = ? WHERE id = ?", (chat_session_id, cot_id))
    except Exception:
        pass

    numero = generar_numero_cotizacion(cot_id)
    execute("UPDATE cotizaciones SET numero = ? WHERE id = ?", (numero, cot_id))

    subtotal_bruto = 0.0
    venta_neta_total = 0.0
    costo_neto_total = 0.0
    contrib_total = 0.0
    items_guardados = 0
    revisar = 0

    for item in items_extraidos:
        descripcion_solicitada = str(item.get("descripcion") or item.get("producto") or "").strip()
        if not descripcion_solicitada or is_noise_line(descripcion_solicitada):
            continue

        codigo_solicitado = item.get("codigo_producto") or item.get("codigo")
        cantidad = parse_quantity(item.get("cantidad"), 1)

        producto_result, score = buscar_producto_local(descripcion_solicitada, codigo_solicitado)

        requiere_revision = 0
        if isinstance(producto_result, dict) and "__candidato__" in producto_result:
            candidato = producto_result["__candidato__"]
            codigo = ""
            descripcion = "REVISAR CANDIDATO: " + candidato["descripcion"]
            compra = 0.0
            venta_bruto = 0.0
            stock = float(candidato["stock"] or 0)
            encontrado = 0
            requiere_revision = 1
            obs = f"Candidato no confirmado. Score {score}. No se suma al total."
            revisar += 1
        elif producto_result:
            producto = producto_result
            codigo = producto["codigo_producto"]
            descripcion = producto["descripcion"]
            compra = float(producto["precio_compra_neto"] or 0)
            venta_bruto = float(producto["precio_venta_bruto"] or 0)
            stock = float(producto["stock"] or 0)
            encontrado = 1
            obs = item.get("observacion") or f"Match confiable score {score}"
        else:
            codigo = str(codigo_solicitado or "")
            descripcion = "PENDIENTE DE IDENTIFICAR"
            compra = 0.0
            venta_bruto = 0.0
            stock = 0.0
            encontrado = 0
            requiere_revision = 1
            obs = f"No encontrado en maestra. Score {score}"
            revisar += 1

        venta_neta_unit = venta_neta_desde_bruto(venta_bruto)
        contrib_unit = venta_neta_unit - compra
        subtotal_item = venta_bruto * cantidad if encontrado else 0.0
        venta_neta_item = venta_neta_unit * cantidad if encontrado else 0.0
        costo_item = compra * cantidad if encontrado else 0.0
        contrib_item = contrib_unit * cantidad if encontrado else 0.0
        margen_item = (contrib_unit / venta_neta_unit) if venta_neta_unit > 0 and encontrado else 0

        execute("""
            INSERT INTO cotizacion_items (
                cotizacion_id, codigo_producto, descripcion_solicitada, descripcion_producto,
                cantidad, precio_compra_neto, precio_venta_bruto, venta_neta_unitaria,
                stock, subtotal_bruto, costo_neto_total, contribucion_total, margen_pct,
                encontrado, observacion, created_at
            )
            VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
        """, (
            cot_id, codigo, descripcion_solicitada, descripcion, cantidad,
            compra, venta_bruto, venta_neta_unit, stock, subtotal_item, costo_item,
            contrib_item, margen_item, encontrado, obs, now
        ))

        # Actualiza columnas nuevas si existen.
        try:
            item_id = query_one("SELECT last_insert_rowid() id")["id"]
            execute("UPDATE cotizacion_items SET match_score=?, requiere_revision=? WHERE id=?", (score or 0, requiere_revision, item_id))
        except Exception:
            pass

        subtotal_bruto += subtotal_item
        venta_neta_total += venta_neta_item
        costo_neto_total += costo_item
        contrib_total += contrib_item
        items_guardados += 1

    margen_total = (contrib_total / venta_neta_total) if venta_neta_total > 0 else 0
    estado = "Requiere revisión" if revisar else "Borrador generado"

    execute("""
        UPDATE cotizaciones
        SET subtotal_bruto=?, venta_neta_total=?, costo_neto_total=?, contribucion_total=?,
            margen_total_pct=?, estado=?, updated_at=?
        WHERE id=?
    """, (subtotal_bruto, venta_neta_total, costo_neto_total, contrib_total, margen_total, estado, now_str(), cot_id))

    write_audit("ventas", "crear_cotizacion", cot_id, "items", "", f"{items_guardados} productos, {revisar} revisar")
    return cot_id


def header_index_map(headers):
    return {normalize_text(h): idx for idx, h in enumerate(headers) if h is not None}


def find_col(index_map, aliases):
    for alias in aliases:
        key = normalize_text(alias)
        if key in index_map:
            return index_map[key]
    return None


def procesar_excel_productos(file_storage):
    from openpyxl import load_workbook

    filename = secure_filename(file_storage.filename or "productos.xlsx")
    backup_database_if_exists("pre_import_productos")

    wb = load_workbook(file_storage, read_only=True, data_only=True)
    ws = wb.active

    rows = ws.iter_rows(values_only=True)
    headers = next(rows, None)
    if not headers:
        raise ValueError("El archivo no tiene encabezados.")

    idx = header_index_map(headers)

    col_codigo = find_col(idx, ["Código Producto", "Codigo Producto", "Codigo", "Código", "SKU"])
    col_desc = find_col(idx, ["Descripción", "Descripcion", "Nombre Producto", "Producto"])
    col_compra = find_col(idx, ["Precio Compra Neto", "Compra Neto", "Costo Neto", "Precio Costo Neto"])
    col_venta_bruto = find_col(idx, ["Precio Venta Bruto", "Venta Bruto", "Precio Bruto", "Precio Venta"])
    col_venta_neto = find_col(idx, ["Precio Venta Neto", "Venta Neto"])
    col_stock = find_col(idx, ["Stock", "Existencia", "Inventario"])
    col_activo = find_col(idx, ["Activo", "Activa", "Estado"])

    missing = []
    if col_codigo is None: missing.append("Código Producto")
    if col_desc is None: missing.append("Descripción")
    if col_compra is None: missing.append("Precio Compra Neto")
    if col_venta_bruto is None and col_venta_neto is None: missing.append("Precio Venta Bruto o Precio Venta Neto")
    if col_stock is None: missing.append("Stock")
    if col_activo is None: missing.append("Activo")

    if missing:
        raise ValueError("Faltan columnas obligatorias: " + ", ".join(missing))

    user = current_user()
    import_id = insert_and_get_id("""
        INSERT INTO producto_importaciones (
            archivo_nombre, total_filas, creados, actualizados, errores,
            usuario_id, usuario_nombre, created_at, observaciones
        )
        VALUES (?, 0, 0, 0, 0, ?, ?, ?, ?)
    """, (
        filename,
        user["id"] if user else None,
        user["full_name"] if user else "Sistema",
        now_str(),
        "Importación iniciada"
    ))

    total = creados = actualizados = errores = 0
    now = now_str()
    conn = db()
    try:
        for row_num, row in enumerate(rows, start=2):
            total += 1
            try:
                codigo = str(row[col_codigo]).strip() if row[col_codigo] is not None else ""
                descripcion = str(row[col_desc]).strip() if row[col_desc] is not None else ""

                if not codigo or not descripcion:
                    raise ValueError("Código Producto y Descripción son obligatorios.")

                compra = parse_number(row[col_compra], 0)
                if col_venta_bruto is not None:
                    venta_bruto = parse_number(row[col_venta_bruto], 0)
                else:
                    venta_neto = parse_number(row[col_venta_neto], 0)
                    venta_bruto = venta_neto * (1 + iva_rate())

                stock = parse_number(row[col_stock], 0)
                activo = parse_bool_si(row[col_activo])

                exists = conn.execute("SELECT id FROM productos WHERE codigo_producto = ?", (codigo,)).fetchone()
                if exists:
                    actualizados += 1
                else:
                    creados += 1

                conn.execute("""
                    INSERT INTO productos (
                        codigo_producto, descripcion, descripcion_busqueda, precio_compra_neto,
                        precio_venta_bruto, stock, activo, ultima_importacion_id, created_at, updated_at
                    )
                    VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
                    ON CONFLICT(codigo_producto) DO UPDATE SET
                        descripcion = excluded.descripcion,
                        descripcion_busqueda = excluded.descripcion_busqueda,
                        precio_compra_neto = excluded.precio_compra_neto,
                        precio_venta_bruto = excluded.precio_venta_bruto,
                        stock = excluded.stock,
                        activo = excluded.activo,
                        ultima_importacion_id = excluded.ultima_importacion_id,
                        updated_at = excluded.updated_at
                """, (
                    codigo, descripcion, normalize_text(descripcion), compra, venta_bruto,
                    stock, activo, import_id, now, now
                ))

            except Exception as exc:
                errores += 1
                conn.execute("""
                    INSERT INTO producto_importacion_errores (
                        importacion_id, fila, codigo_producto, error, created_at
                    )
                    VALUES (?, ?, ?, ?, ?)
                """, (
                    import_id,
                    row_num,
                    "" if not row or col_codigo is None or col_codigo >= len(row) else str(row[col_codigo] or ""),
                    str(exc)[:500],
                    now_str()
                ))

        conn.execute("""
            UPDATE producto_importaciones
            SET total_filas=?, creados=?, actualizados=?, errores=?, observaciones=?
            WHERE id=?
        """, (total, creados, actualizados, errores, "Importación finalizada", import_id))
        conn.commit()
    finally:
        conn.close()

    write_audit("productos", "importar_excel", import_id, "archivo", "", f"{filename}: {total} filas")
    return {
        "import_id": import_id,
        "archivo": filename,
        "total": total,
        "creados": creados,
        "actualizados": actualizados,
        "errores": errores,
    }


# ============================================================
# INICIALIZACIÓN / MIGRACIONES
# ============================================================

def init_db():
    # Respaldo automático antes de cualquier migración. No elimina datos.
    backup_database_if_exists('pre_migracion')
    execute("""
    CREATE TABLE IF NOT EXISTS users (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        username TEXT UNIQUE NOT NULL,
        full_name TEXT NOT NULL,
        password_hash TEXT NOT NULL,
        role TEXT NOT NULL DEFAULT 'operador',
        is_active INTEGER NOT NULL DEFAULT 1,
        permissions TEXT NOT NULL DEFAULT '{}',
        created_at TEXT NOT NULL,
        updated_at TEXT
    )
    """)

    execute("""
    CREATE TABLE IF NOT EXISTS despachos (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        numero_documento TEXT NOT NULL,
        tipo_documento TEXT NOT NULL,
        estado TEXT NOT NULL,
        cliente TEXT NOT NULL,
        telefono TEXT,
        destino TEXT,
        patente TEXT,
        conductor TEXT,
        pioneta TEXT,
        monto REAL DEFAULT 0,
        usuario_id INTEGER,
        usuario_nombre TEXT,
        created_at TEXT NOT NULL,
        updated_at TEXT
    )
    """)

    execute("""
    CREATE TABLE IF NOT EXISTS maquinarias (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        codigo TEXT,
        nombre TEXT NOT NULL,
        tipo TEXT,
        marca TEXT,
        modelo TEXT,
        anio TEXT,
        patente TEXT,
        estado TEXT NOT NULL DEFAULT 'Activa',
        observaciones TEXT,
        created_at TEXT NOT NULL,
        updated_at TEXT
    )
    """)

    execute("""
    CREATE TABLE IF NOT EXISTS mantenciones (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        maquinaria_id INTEGER,
        maquinaria_nombre TEXT,
        fecha TEXT NOT NULL,
        tipo_mantencion TEXT NOT NULL,
        estado TEXT NOT NULL DEFAULT 'Pendiente',
        responsable TEXT,
        costo REAL DEFAULT 0,
        observaciones TEXT,
        usuario_id INTEGER,
        usuario_nombre TEXT,
        created_at TEXT NOT NULL,
        updated_at TEXT
    )
    """)

    execute("""
    CREATE TABLE IF NOT EXISTS vehiculos (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        patente TEXT UNIQUE NOT NULL,
        tipo TEXT,
        marca TEXT,
        modelo TEXT,
        anio TEXT,
        estado TEXT NOT NULL DEFAULT 'Activo',
        permiso_circulacion_vencimiento TEXT,
        revision_tecnica_vencimiento TEXT,
        seguro_obligatorio_vencimiento TEXT,
        observaciones TEXT,
        created_at TEXT NOT NULL,
        updated_at TEXT
    )
    """)

    execute("""
    CREATE TABLE IF NOT EXISTS personas_logistica (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        nombre TEXT NOT NULL,
        tipo TEXT NOT NULL,
        estado TEXT NOT NULL DEFAULT 'Activo',
        telefono TEXT,
        observaciones TEXT,
        created_at TEXT NOT NULL,
        updated_at TEXT
    )
    """)

    execute("""
    CREATE TABLE IF NOT EXISTS audit_log (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        modulo TEXT NOT NULL,
        accion TEXT NOT NULL,
        registro_id INTEGER,
        campo TEXT,
        valor_anterior TEXT,
        valor_nuevo TEXT,
        usuario_id INTEGER,
        usuario_nombre TEXT,
        created_at TEXT NOT NULL
    )
    """)

    execute("""
    CREATE TABLE IF NOT EXISTS system_config (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        clave TEXT UNIQUE NOT NULL,
        valor TEXT NOT NULL,
        updated_at TEXT
    )
    """)


    execute("""
    CREATE TABLE IF NOT EXISTS productos (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        codigo_producto TEXT UNIQUE NOT NULL,
        descripcion TEXT NOT NULL,
        descripcion_busqueda TEXT,
        precio_compra_neto REAL DEFAULT 0,
        precio_venta_bruto REAL DEFAULT 0,
        stock REAL DEFAULT 0,
        activo INTEGER DEFAULT 1,
        ultima_importacion_id INTEGER,
        created_at TEXT NOT NULL,
        updated_at TEXT
    )
    """)

    execute("""
    CREATE TABLE IF NOT EXISTS producto_importaciones (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        archivo_nombre TEXT,
        total_filas INTEGER DEFAULT 0,
        creados INTEGER DEFAULT 0,
        actualizados INTEGER DEFAULT 0,
        errores INTEGER DEFAULT 0,
        usuario_id INTEGER,
        usuario_nombre TEXT,
        created_at TEXT NOT NULL,
        observaciones TEXT
    )
    """)

    execute("""
    CREATE TABLE IF NOT EXISTS producto_importacion_errores (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        importacion_id INTEGER,
        fila INTEGER,
        codigo_producto TEXT,
        error TEXT,
        created_at TEXT NOT NULL
    )
    """)

    execute("""
    CREATE TABLE IF NOT EXISTS cotizaciones (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        numero TEXT UNIQUE,
        cliente TEXT,
        telefono TEXT,
        texto_original TEXT,
        fuente TEXT,
        ai_raw TEXT,
        subtotal_bruto REAL DEFAULT 0,
        venta_neta_total REAL DEFAULT 0,
        costo_neto_total REAL DEFAULT 0,
        contribucion_total REAL DEFAULT 0,
        margen_total_pct REAL DEFAULT 0,
        estado TEXT DEFAULT 'Generada',
        usuario_id INTEGER,
        usuario_nombre TEXT,
        created_at TEXT NOT NULL,
        updated_at TEXT
    )
    """)

    execute("""
    CREATE TABLE IF NOT EXISTS cotizacion_items (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        cotizacion_id INTEGER NOT NULL,
        codigo_producto TEXT,
        descripcion_solicitada TEXT,
        descripcion_producto TEXT,
        cantidad REAL DEFAULT 1,
        precio_compra_neto REAL DEFAULT 0,
        precio_venta_bruto REAL DEFAULT 0,
        venta_neta_unitaria REAL DEFAULT 0,
        stock REAL DEFAULT 0,
        subtotal_bruto REAL DEFAULT 0,
        costo_neto_total REAL DEFAULT 0,
        contribucion_total REAL DEFAULT 0,
        margen_pct REAL DEFAULT 0,
        encontrado INTEGER DEFAULT 0,
        observacion TEXT,
        created_at TEXT NOT NULL,
        FOREIGN KEY(cotizacion_id) REFERENCES cotizaciones(id)
    )
    """)

    execute("""
    CREATE TABLE IF NOT EXISTS ventas_chat_sesiones (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        titulo TEXT,
        cliente TEXT,
        telefono TEXT,
        estado TEXT DEFAULT 'Abierta',
        usuario_id INTEGER,
        usuario_nombre TEXT,
        created_at TEXT NOT NULL,
        updated_at TEXT
    )
    """)

    execute("""
    CREATE TABLE IF NOT EXISTS ventas_chat_mensajes (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        sesion_id INTEGER NOT NULL,
        rol TEXT NOT NULL,
        contenido TEXT NOT NULL,
        ai_raw TEXT,
        tiene_imagen INTEGER DEFAULT 0,
        created_at TEXT NOT NULL,
        FOREIGN KEY(sesion_id) REFERENCES ventas_chat_sesiones(id)
    )
    """)

    execute("CREATE INDEX IF NOT EXISTS idx_ventas_chat_mensajes_sesion ON ventas_chat_mensajes(sesion_id)")
    execute("CREATE INDEX IF NOT EXISTS idx_ventas_chat_sesiones_updated ON ventas_chat_sesiones(updated_at)")

    execute("CREATE INDEX IF NOT EXISTS idx_productos_descripcion_busqueda ON productos(descripcion_busqueda)")
    execute("CREATE INDEX IF NOT EXISTS idx_cotizaciones_created_at ON cotizaciones(created_at)")
    execute("CREATE INDEX IF NOT EXISTS idx_cotizacion_items_cotizacion ON cotizacion_items(cotizacion_id)")

    # Migraciones tolerantes para bases anteriores.
    migrations = {
        "users": [
            "full_name TEXT DEFAULT ''",
            "password_hash TEXT DEFAULT ''",
            "role TEXT DEFAULT 'operador'",
            "is_active INTEGER DEFAULT 1",
            "permissions TEXT DEFAULT '{}'",
            "created_at TEXT DEFAULT ''",
            "updated_at TEXT",
        ],
        "despachos": [
            "numero_documento TEXT DEFAULT ''",
            "tipo_documento TEXT DEFAULT 'Factura'",
            "estado TEXT DEFAULT 'Pendiente'",
            "cliente TEXT DEFAULT ''",
            "telefono TEXT",
            "destino TEXT",
            "patente TEXT",
            "conductor TEXT",
            "pioneta TEXT",
            "monto REAL DEFAULT 0",
            "usuario_id INTEGER",
            "usuario_nombre TEXT",
            "created_at TEXT DEFAULT ''",
            "updated_at TEXT",
        ],
        "maquinarias": [
            "codigo TEXT",
            "nombre TEXT DEFAULT ''",
            "tipo TEXT",
            "marca TEXT",
            "modelo TEXT",
            "anio TEXT",
            "patente TEXT",
            "estado TEXT DEFAULT 'Activa'",
            "observaciones TEXT",
            "created_at TEXT DEFAULT ''",
            "updated_at TEXT",
        ],
        "mantenciones": [
            "maquinaria_id INTEGER",
            "maquinaria_nombre TEXT",
            "fecha TEXT DEFAULT ''",
            "tipo_mantencion TEXT DEFAULT 'Preventiva'",
            "estado TEXT DEFAULT 'Pendiente'",
            "responsable TEXT",
            "costo REAL DEFAULT 0",
            "observaciones TEXT",
            "usuario_id INTEGER",
            "usuario_nombre TEXT",
            "created_at TEXT DEFAULT ''",
            "updated_at TEXT",
        ],
    }

    for table, cols in migrations.items():
        for col in cols:
            try:
                add_column_if_missing(table, col)
            except Exception:
                pass

    # Migraciones de módulos Ventas / Productos.
    extra_migration_cols = {
        "productos": [
            "descripcion_busqueda TEXT",
            "ultima_importacion_id INTEGER",
        ],
        "cotizaciones": [
            "fuente TEXT",
            "ai_raw TEXT",
            "venta_neta_total REAL DEFAULT 0",
            "costo_neto_total REAL DEFAULT 0",
            "contribucion_total REAL DEFAULT 0",
            "margen_total_pct REAL DEFAULT 0",
            "chat_session_id INTEGER",
        ],
        "cotizacion_items": [
            "venta_neta_unitaria REAL DEFAULT 0",
            "descripcion_solicitada TEXT",
            "match_score REAL DEFAULT 0",
            "requiere_revision INTEGER DEFAULT 0",
        ],
        "ventas_chat_sesiones": [
            "cliente TEXT",
            "telefono TEXT",
            "estado TEXT DEFAULT 'Abierta'",
            "usuario_id INTEGER",
            "usuario_nombre TEXT",
            "created_at TEXT DEFAULT ''",
            "updated_at TEXT",
        ],
        "ventas_chat_mensajes": [
            "ai_raw TEXT",
            "tiene_imagen INTEGER DEFAULT 0",
        ],
    }
    for table, cols in extra_migration_cols.items():
        for col in cols:
            try:
                add_column_if_missing(table, col)
            except Exception:
                pass

    # Configuración inicial.
    defaults = {
        "estados_despacho": ["Entregado", "Pendiente"],
        "tipos_documento": ["Factura", "Boleta", "Guía de despacho", "Otro"],
        "sucursales": ["Sucursal 1", "Sucursal 2", "La Americana"],
        "estados_maquinaria": ["Activa", "En mantención", "Fuera de servicio", "Vendida / dada de baja"],
        "estados_mantencion": ["Pendiente", "En proceso", "Realizada", "Anulada"],
        "tipos_mantencion": ["Preventiva", "Correctiva", "Emergencia", "Revisión", "Otro"],
    }
    for clave, valor in defaults.items():
        row = query_one("SELECT id FROM system_config WHERE clave = ?", (clave,))
        if not row:
            set_config_list(clave, valor)

    # No se sobreescribe configuración existente para evitar pérdida de ajustes.
    # El formulario rápido de despacho usa solo Entregado/Pendiente en la interfaz.

    # Usuario admin inicial.
    admin = query_one("SELECT id FROM users WHERE username = ?", ("admin",))
    if not admin:
        execute("""
            INSERT INTO users (
                username, full_name, password_hash, role, is_active,
                permissions, created_at, updated_at
            )
            VALUES (?, ?, ?, ?, ?, ?, ?, ?)
        """, (
            "admin",
            "Administrador",
            generate_password_hash("admin123"),
            "admin",
            1,
            json.dumps({"all": True}),
            now_str(),
            now_str(),
        ))

    # Usuario operativo inicial de referencia.
    operador = query_one("SELECT id FROM users WHERE username = ?", ("operador",))
    if not operador:
        execute("""
            INSERT INTO users (
                username, full_name, password_hash, role, is_active,
                permissions, created_at, updated_at
            )
            VALUES (?, ?, ?, ?, ?, ?, ?, ?)
        """, (
            "operador",
            "Usuario Operador",
            generate_password_hash("operador123"),
            "operador",
            1,
            json.dumps({
                "despachos": True,
                "mantenciones": True,
                "consulta": True,
                "ventas": True,
                "productos": False,
                "exportar": False,
                "dashboard": False,
                "auditoria": False,
                "administracion": False,
            }),
            now_str(),
            now_str(),
        ))


# ============================================================
# SEGURIDAD / SESIÓN
# ============================================================

PERMISSION_LABELS = {
    "ventas": "Ventas / Cotización IA",
    "productos": "Productos / Maestro de precios",
    "despachos": "Despachos",
    "mantenciones": "Mantenciones",
    "consulta": "Consulta",
    "dashboard": "Dashboard admin",
    "auditoria": "Auditoría",
    "administracion": "Administración / Configuración",
    "usuarios": "Usuarios",
    "maquinarias": "Maquinarias",
    "vehiculos": "Vehículos / Patentes",
    "logistica": "Conductores / Pionetas",
    "exportar": "Exportar Excel",
    "integraciones": "Integraciones",
}


def current_user():
    uid = session.get("user_id")
    if not uid:
        return None
    return query_one("SELECT * FROM users WHERE id = ? AND is_active = 1", (uid,))


def user_permissions(user=None):
    user = user or current_user()
    if not user:
        return {}
    if user["role"] == "admin":
        return {"all": True}
    try:
        return json.loads(user["permissions"] or "{}")
    except Exception:
        return {}


def is_admin(user=None):
    user = user or current_user()
    return bool(user and user["role"] == "admin")


def has_perm(module_name):
    user = current_user()
    if not user:
        return False
    if is_admin(user):
        return True
    perms = user_permissions(user)
    return bool(perms.get(module_name))


def login_required(fn):
    @wraps(fn)
    def wrapper(*args, **kwargs):
        if not current_user():
            return redirect(url_for("login"))
        return fn(*args, **kwargs)
    return wrapper


def permission_required(module_name):
    def decorator(fn):
        @wraps(fn)
        def wrapper(*args, **kwargs):
            if not current_user():
                return redirect(url_for("login"))
            if not has_perm(module_name):
                flash("No tienes permiso para acceder a esa sección.", "error")
                return redirect(url_for("index"))
            return fn(*args, **kwargs)
        return wrapper
    return decorator


def write_audit(modulo, accion, registro_id=None, campo=None, valor_anterior=None, valor_nuevo=None):
    user = current_user()
    execute("""
        INSERT INTO audit_log (
            modulo, accion, registro_id, campo, valor_anterior, valor_nuevo,
            usuario_id, usuario_nombre, created_at
        )
        VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?)
    """, (
        modulo,
        accion,
        registro_id,
        campo,
        "" if valor_anterior is None else str(valor_anterior),
        "" if valor_nuevo is None else str(valor_nuevo),
        user["id"] if user else None,
        user["full_name"] if user else "Sistema",
        now_str(),
    ))


# ============================================================
# HTML BASE
# ============================================================

BASE_TEMPLATE = r"""
<!doctype html>
<html lang="es">
<head>
    <meta charset="utf-8">
    <title>{{ page_title }} · {{ app_name }}</title>
    <meta name="viewport" content="width=device-width, initial-scale=1">
    <style>
        :root{
            --bg:#f5f7fb;
            --panel:#ffffff;
            --panel-2:#f9fafb;
            --text:#111827;
            --muted:#6b7280;
            --border:#e5e7eb;
            --primary:#0f766e;
            --primary-dark:#115e59;
            --danger:#dc2626;
            --warning:#d97706;
            --ok:#059669;
            --shadow:0 10px 30px rgba(15,23,42,.08);
            --radius:18px;
        }
        *{box-sizing:border-box}
        body{
            margin:0;
            font-family:Inter,Segoe UI,Roboto,Arial,sans-serif;
            background:var(--bg);
            color:var(--text);
        }
        a{color:inherit;text-decoration:none}
        .login-bg{
            min-height:100vh;
            display:flex;
            align-items:center;
            justify-content:center;
            padding:24px;
            background:
                radial-gradient(circle at 18% 18%, rgba(20,184,166,.22), transparent 28%),
                radial-gradient(circle at 80% 72%, rgba(59,130,246,.24), transparent 30%),
                linear-gradient(135deg,#071827 0%, #0f2437 50%, #0e5962 100%);
            position:relative;
            overflow:hidden;
        }
        .login-bg::before{
            content:"";
            position:absolute;
            inset:0;
            background:
                linear-gradient(rgba(255,255,255,.04) 1px, transparent 1px),
                linear-gradient(90deg, rgba(255,255,255,.04) 1px, transparent 1px);
            background-size:34px 34px;
            mask-image:linear-gradient(to bottom, rgba(0,0,0,.75), transparent 95%);
        }
        .login-card{
            width:100%;
            max-width:430px;
            background:rgba(255,255,255,.92);
            border:1px solid rgba(255,255,255,.55);
            border-radius:28px;
            padding:34px 30px 24px;
            box-shadow:0 24px 70px rgba(2,6,23,.42);
            backdrop-filter:blur(18px);
            position:relative;
            z-index:1;
        }
        .topbar{
            background:#0f172a;
            color:white;
            padding:14px 20px;
            display:flex;
            align-items:center;
            justify-content:space-between;
            gap:16px;
            position:sticky;
            top:0;
            z-index:20;
            box-shadow:0 4px 18px rgba(0,0,0,.15);
        }
        .brand{
            display:flex;
            flex-direction:column;
            gap:2px;
            min-width:230px;
        }
        .brand strong{font-size:16px}
        .brand small{font-size:12px;color:#cbd5e1}
        .nav-wrap{
            display:flex;
            flex-wrap:wrap;
            justify-content:center;
            gap:10px;
            flex:1;
        }
        .nav-section{
            display:flex;
            align-items:center;
            gap:6px;
            padding:6px;
            border-radius:14px;
            border:1px solid rgba(255,255,255,.14);
            background:rgba(255,255,255,.07);
        }
        .nav-section.admin{background:rgba(20,184,166,.18);border-color:rgba(45,212,191,.35)}
        .nav-section.control{background:rgba(251,191,36,.15);border-color:rgba(251,191,36,.35)}
        .nav-section.integrations{background:rgba(129,140,248,.17);border-color:rgba(129,140,248,.35)}
        .nav-section.ventas{background:rgba(14,165,233,.16);border-color:rgba(56,189,248,.35)}
        .nav-label{
            color:#cbd5e1;
            font-size:11px;
            font-weight:700;
            text-transform:uppercase;
            letter-spacing:.06em;
            padding:0 4px;
        }
        .nav-link{
            font-size:13px;
            padding:8px 10px;
            border-radius:10px;
            color:#f8fafc;
        }
        .nav-link:hover{background:rgba(255,255,255,.14)}
        .userbox{
            display:flex;
            align-items:center;
            gap:8px;
            font-size:13px;
            color:#e5e7eb;
        }
        .logout{
            background:#334155;
            color:white;
            padding:8px 10px;
            border-radius:10px;
        }
        .layout{max-width:1450px;margin:0 auto;padding:22px}
        .page-head{
            display:flex;
            align-items:flex-start;
            justify-content:space-between;
            gap:16px;
            margin-bottom:18px;
        }
        .page-head h1{margin:0;font-size:28px}
        .page-head p{margin:6px 0 0;color:var(--muted)}
        .card{
            background:var(--panel);
            border:1px solid var(--border);
            border-radius:var(--radius);
            box-shadow:var(--shadow);
            padding:20px;
            margin-bottom:18px;
        }
        .card h2,.card h3{margin-top:0}
        .grid{display:grid;gap:16px}
        .grid-2{grid-template-columns:repeat(2,minmax(0,1fr))}
        .grid-3{grid-template-columns:repeat(3,minmax(0,1fr))}
        .grid-4{grid-template-columns:repeat(4,minmax(0,1fr))}
        .stat{
            background:white;
            border:1px solid var(--border);
            border-radius:18px;
            padding:18px;
            box-shadow:0 6px 18px rgba(15,23,42,.05);
        }
        .stat span{display:block;color:var(--muted);font-size:13px}
        .stat strong{display:block;font-size:30px;margin-top:8px}
        label{
            display:block;
            font-size:13px;
            font-weight:700;
            margin-bottom:6px;
            color:#374151;
        }
        input,select,textarea{
            width:100%;
            border:1px solid #d1d5db;
            border-radius:12px;
            padding:11px 12px;
            font-size:14px;
            background:white;
            color:#111827;
        }
        textarea{min-height:96px;resize:vertical}
        .form-row{display:grid;grid-template-columns:repeat(4,minmax(0,1fr));gap:14px}
        .form-row-3{display:grid;grid-template-columns:repeat(3,minmax(0,1fr));gap:14px}
        .form-row-2{display:grid;grid-template-columns:repeat(2,minmax(0,1fr));gap:14px}
        .actions{display:flex;gap:10px;align-items:center;flex-wrap:wrap;margin-top:16px}
        .btn{
            border:0;
            border-radius:12px;
            padding:11px 14px;
            font-weight:800;
            cursor:pointer;
            display:inline-flex;
            align-items:center;
            gap:8px;
            font-size:14px;
        }
        .btn-primary{background:var(--primary);color:white}
        .btn-primary:hover{background:var(--primary-dark)}
        .btn-secondary{background:#e5e7eb;color:#111827}
        .btn-danger{background:var(--danger);color:white}
        .btn-warning{background:var(--warning);color:white}
        .btn-small{padding:7px 9px;border-radius:9px;font-size:12px}
        table{
            width:100%;
            border-collapse:separate;
            border-spacing:0;
            overflow:hidden;
            border:1px solid var(--border);
            border-radius:16px;
            background:white;
        }
        th,td{
            padding:11px 12px;
            border-bottom:1px solid var(--border);
            text-align:left;
            font-size:13px;
            vertical-align:top;
        }
        th{
            background:#f8fafc;
            color:#475569;
            font-size:12px;
            text-transform:uppercase;
            letter-spacing:.04em;
        }
        tr:last-child td{border-bottom:0}
        .table-wrap{overflow:auto}
        .badge{
            display:inline-flex;
            align-items:center;
            border-radius:999px;
            padding:4px 9px;
            font-weight:800;
            font-size:12px;
            background:#e5e7eb;
            color:#374151;
            white-space:nowrap;
        }
        .badge.ok{background:#d1fae5;color:#065f46}
        .badge.warn{background:#fef3c7;color:#92400e}
        .badge.bad{background:#fee2e2;color:#991b1b}
        .badge.info{background:#dbeafe;color:#1e40af}
        .flash{
            padding:12px 14px;
            border-radius:14px;
            margin-bottom:12px;
            font-weight:700;
        }
        .flash.success{background:#d1fae5;color:#065f46}
        .flash.error{background:#fee2e2;color:#991b1b}
        .flash.info{background:#dbeafe;color:#1e40af}
        .muted{color:var(--muted)}
        .section-title{
            display:flex;
            align-items:center;
            justify-content:space-between;
            gap:12px;
            margin-bottom:12px;
        }
        .danger-zone{
            border:1px solid #fecaca;
            background:#fff7f7;
        }
        .checkbox-grid{
            display:grid;
            grid-template-columns:repeat(3,minmax(0,1fr));
            gap:10px;
        }
        .check-item{
            display:flex;
            align-items:center;
            gap:8px;
            border:1px solid var(--border);
            border-radius:12px;
            padding:10px;
            background:#fff;
        }
        .check-item input{width:auto}
        .placeholder{
            border:2px dashed #cbd5e1;
            background:#f8fafc;
            border-radius:18px;
            padding:20px;
            color:#475569;
        }
        .ai-panel{
            background:
                radial-gradient(circle at top left, rgba(20,184,166,.18), transparent 24%),
                linear-gradient(135deg,#071827,#0f172a);
            color:#e5f8ff;
            border:1px solid rgba(125,211,252,.25);
            border-radius:22px;
            padding:20px;
            box-shadow:0 18px 50px rgba(15,23,42,.18);
        }
        .ai-panel label{color:#cbd5e1}
        .ai-panel input,.ai-panel textarea{
            background:rgba(255,255,255,.07);
            color:#f8fafc;
            border-color:rgba(148,163,184,.30);
        }
        .ai-panel textarea{min-height:160px}
        .subtle{
            color:#94a3b8;
            font-size:13px;
        }
        .quote-total{
            background:#f8fafc;
            border:1px solid var(--border);
            border-radius:16px;
            padding:14px;
        }
        .match-ok{color:#047857;font-weight:800}
        .match-review{color:#b45309;font-weight:800}
        .match-bad{color:#b91c1c;font-weight:800}
        .sales-layout{
            display:grid;
            grid-template-columns:minmax(0,1fr) 360px;
            gap:18px;
            align-items:start;
        }
        .chat-shell{
            min-height:680px;
            display:flex;
            flex-direction:column;
        }
        .chat-box{
            background:
                radial-gradient(circle at top right, rgba(37,99,235,.15), transparent 24%),
                rgba(255,255,255,.045);
            border:1px solid rgba(148,163,184,.25);
            border-radius:20px;
            padding:18px;
            height:560px;
            overflow:auto;
            display:flex;
            flex-direction:column;
            gap:14px;
            scroll-behavior:smooth;
        }
        .chat-msg{
            max-width:82%;
            border-radius:18px;
            padding:12px 14px;
            white-space:pre-wrap;
            line-height:1.45;
            font-size:14px;
            box-shadow:0 10px 24px rgba(2,6,23,.16);
        }
        .chat-msg.user{
            align-self:flex-end;
            background:linear-gradient(135deg, rgba(20,184,166,.35), rgba(14,116,144,.35));
            border:1px solid rgba(45,212,191,.38);
            color:#ecfeff;
            border-bottom-right-radius:6px;
        }
        .chat-msg.assistant{
            align-self:flex-start;
            background:rgba(15,23,42,.72);
            border:1px solid rgba(96,165,250,.30);
            color:#e5eefb;
            border-bottom-left-radius:6px;
        }
        .chat-msg.quote{
            align-self:stretch;
            max-width:100%;
            background:#ffffff;
            color:#0f172a;
            border:1px solid #dbeafe;
            border-radius:20px;
            padding:0;
            overflow:hidden;
            white-space:normal;
        }
        .quote-chat-head{
            background:linear-gradient(135deg,#0f766e,#2563eb);
            color:white;
            padding:14px 16px;
            display:flex;
            justify-content:space-between;
            gap:12px;
            align-items:flex-start;
        }
        .quote-chat-body{padding:14px 16px}
        .quote-chat-totals{
            display:grid;
            grid-template-columns:repeat(4,minmax(0,1fr));
            gap:10px;
            margin:12px 0;
        }
        .quote-chip{
            border:1px solid #e5e7eb;
            border-radius:14px;
            padding:10px;
            background:#f8fafc;
        }
        .quote-chip span{display:block;color:#64748b;font-size:11px}
        .quote-chip strong{display:block;margin-top:4px;font-size:15px}
        .quote-table-mini{
            width:100%;
            border-collapse:collapse;
            border:1px solid #e5e7eb;
            border-radius:12px;
            overflow:hidden;
            font-size:12px;
        }
        .quote-table-mini th,.quote-table-mini td{
            padding:8px;
            border-bottom:1px solid #e5e7eb;
            vertical-align:top;
        }
        .chat-compose{
            border:1px solid rgba(148,163,184,.25);
            background:rgba(15,23,42,.58);
            border-radius:20px;
            padding:14px;
            margin-top:14px;
        }
        .chat-compose textarea{
            min-height:92px;
            max-height:180px;
        }
        .side-card{
            background:white;
            border:1px solid var(--border);
            border-radius:18px;
            padding:18px;
            box-shadow:var(--shadow);
            position:sticky;
            top:90px;
        }
        .confidence{
            display:inline-flex;
            align-items:center;
            border-radius:999px;
            padding:3px 8px;
            font-size:11px;
            font-weight:800;
            background:#e5e7eb;
            color:#374151;
        }
        .confidence.ok{background:#d1fae5;color:#065f46}
        .confidence.review{background:#fef3c7;color:#92400e}
        .confidence.bad{background:#fee2e2;color:#991b1b}
        @media (max-width:1000px){
            .form-row,.form-row-3,.form-row-2,.grid-2,.grid-3,.grid-4{grid-template-columns:1fr}
            .topbar{align-items:flex-start;flex-direction:column}
            .nav-wrap{justify-content:flex-start}
            .brand{min-width:auto}
            .checkbox-grid{grid-template-columns:1fr}
        }
    </style>
</head>
<body>
{% if login_screen %}
    {{ body|safe }}
{% else %}
    <div class="topbar">
        <div class="brand">
            <strong>{{ app_name }}</strong>
            <small>{{ app_version }}</small>
        </div>

        <div class="nav-wrap">
            <div class="nav-section">
                <span class="nav-label">Operación</span>
                {% if has_perm("despachos") %}<a class="nav-link" href="{{ url_for('despachos') }}">Despachos</a>{% endif %}
                {% if has_perm("mantenciones") %}<a class="nav-link" href="{{ url_for('mantenciones') }}">Mantenciones</a>{% endif %}
                {% if has_perm("consulta") %}<a class="nav-link" href="{{ url_for('consulta') }}">Consulta</a>{% endif %}
            </div>

            {% if has_perm("ventas") or has_perm("productos") %}
            <div class="nav-section ventas">
                <span class="nav-label">Ventas</span>
                {% if has_perm("ventas") %}<a class="nav-link" href="{{ url_for('ventas') }}">Cotización Elias</a>{% endif %}
                {% if has_perm("productos") %}<a class="nav-link" href="{{ url_for('productos') }}">Productos</a>{% endif %}
            </div>
            {% endif %}

            {% if is_admin %}
            <div class="nav-section admin">
                <span class="nav-label">Administración</span>
                <a class="nav-link" href="{{ url_for('administracion') }}">Administración</a>
                <a class="nav-link" href="{{ url_for('usuarios') }}">Usuarios</a>
                <a class="nav-link" href="{{ url_for('productos') }}">Productos</a>
                <a class="nav-link" href="{{ url_for('maquinarias') }}">Maquinarias</a>
                <a class="nav-link" href="{{ url_for('vehiculos') }}">Vehículos</a>
                <a class="nav-link" href="{{ url_for('logistica') }}">Conductores</a>
            </div>

            <div class="nav-section control">
                <span class="nav-label">Control</span>
                <a class="nav-link" href="{{ url_for('dashboard') }}">Dashboard</a>
                <a class="nav-link" href="{{ url_for('auditoria') }}">Auditoría</a>
            </div>

            <div class="nav-section integrations">
                <span class="nav-label">Integraciones</span>
                <a class="nav-link" href="{{ url_for('facturacion') }}">Facturación.cl</a>
                <a class="nav-link" href="{{ url_for('exportar') }}">Exportar Excel</a>
            </div>
            {% endif %}
        </div>

        <div class="userbox">
            <span>{{ user.full_name }} · {{ user.role }}</span>
            <a class="logout" href="{{ url_for('logout') }}">Salir</a>
        </div>
    </div>

    <main class="layout">
        {% with messages = get_flashed_messages(with_categories=true) %}
            {% if messages %}
                {% for category, message in messages %}
                    <div class="flash {{ category }}">{{ message }}</div>
                {% endfor %}
            {% endif %}
        {% endwith %}
        {{ body|safe }}
    </main>
{% endif %}
</body>
</html>
"""


def render_page(title, body_template, login_screen=False, **context):
    body = render_template_string(body_template, **context)
    return render_template_string(
        BASE_TEMPLATE,
        page_title=title,
        app_name=APP_NAME,
        app_version=APP_VERSION,
        body=Markup(body),
        login_screen=login_screen,
        user=current_user(),
        is_admin=is_admin(),
        has_perm=has_perm,
    )


# ============================================================
# RUTAS PRINCIPALES
# ============================================================

@app.route("/health")
def health():
    return {"status": "ok", "app": APP_NAME, "version": APP_VERSION}


@app.route("/debug-db")
@login_required
@permission_required("administracion")
def debug_db():
    parent = os.path.dirname(os.path.abspath(DB_PATH))
    return {
        "DATABASE_PATH": DB_PATH,
        "absolute_path": os.path.abspath(DB_PATH),
        "parent_directory": parent,
        "parent_exists": os.path.exists(parent),
        "parent_writable": os.access(parent, os.W_OK) if os.path.exists(parent) else False,
        "database_exists": os.path.exists(DB_PATH),
        "app_version": APP_VERSION,
    }


@app.route("/")
@login_required
def index():
    if is_admin():
        return redirect(url_for("dashboard"))
    if has_perm("despachos"):
        return redirect(url_for("despachos"))
    if has_perm("consulta"):
        return redirect(url_for("consulta"))
    if has_perm("mantenciones"):
        return redirect(url_for("mantenciones"))
    flash("Tu usuario no tiene módulos asignados. Solicita acceso a un administrador.", "error")
    return redirect(url_for("logout"))


@app.route("/login", methods=["GET", "POST"])
def login():
    if request.method == "POST":
        username = request.form.get("username", "").strip()
        password = request.form.get("password", "")

        user = query_one("SELECT * FROM users WHERE username = ? AND is_active = 1", (username,))
        if user and check_password_hash(user["password_hash"], password):
            session.clear()
            session["user_id"] = user["id"]
            write_audit("login", "ingreso", user["id"], "usuario", "", username)
            return redirect(url_for("index"))

        flash("Usuario o contraseña incorrectos.", "error")

    return render_page("Ingreso", r"""
    <div class="login-bg">
        <div class="login-card">
            <h1 style="margin:0 0 22px;color:#0f172a;font-size:34px;line-height:1.05;letter-spacing:-0.03em;">Ferretería Cloud Tool</h1>

            {% with messages = get_flashed_messages(with_categories=true) %}
                {% if messages %}
                    {% for category, message in messages %}
                        <div class="flash {{ category }}">{{ message }}</div>
                    {% endfor %}
                {% endif %}
            {% endwith %}

            <form method="post">
                <div style="margin-bottom:14px;">
                    <label>Usuario</label>
                    <input name="username" autocomplete="username" required>
                </div>
                <div style="margin-bottom:16px;">
                    <label>Contraseña</label>
                    <input type="password" name="password" autocomplete="current-password" required>
                </div>
                <button class="btn btn-primary" style="width:100%;justify-content:center;height:52px;font-size:16px;">Ingresar</button>
            </form>

            <div style="margin-top:18px;text-align:center;color:#64748b;font-size:12px;letter-spacing:.12em;text-transform:uppercase;">
                RUZ AI Systems
            </div>
        </div>
    </div>
    """, login_screen=True)


@app.route("/logout")
def logout():
    session.clear()
    return redirect(url_for("login"))


# ============================================================
# DASHBOARD ADMIN
# ============================================================

@app.route("/dashboard")
@login_required
@permission_required("dashboard")
def dashboard():
    stats = {
        "total_despachos": query_one("SELECT COUNT(*) c FROM despachos")["c"],
        "pendientes": query_one("SELECT COUNT(*) c FROM despachos WHERE estado = 'Pendiente'")["c"],
        "entregados": query_one("SELECT COUNT(*) c FROM despachos WHERE estado = 'Entregado'")["c"],
        "monto_total": query_one("SELECT COALESCE(SUM(monto),0) total FROM despachos")["total"],
        "mant_pendientes": query_one("SELECT COUNT(*) c FROM mantenciones WHERE estado IN ('Pendiente','En proceso')")["c"],
        "maq_fuera": query_one("SELECT COUNT(*) c FROM maquinarias WHERE estado IN ('En mantención','Fuera de servicio')")["c"],
        "usuarios": query_one("SELECT COUNT(*) c FROM users WHERE is_active = 1")["c"],
    }

    por_usuario = query_all("""
        SELECT COALESCE(usuario_nombre, 'Sin usuario') usuario, COUNT(*) total, COALESCE(SUM(monto),0) monto
        FROM despachos
        GROUP BY COALESCE(usuario_nombre, 'Sin usuario')
        ORDER BY total DESC
        LIMIT 10
    """)

    por_estado = query_all("""
        SELECT COALESCE(NULLIF(estado,''), 'Sin estado') estado, COUNT(*) total, COALESCE(SUM(monto),0) monto
        FROM despachos
        GROUP BY COALESCE(NULLIF(estado,''), 'Sin estado')
        ORDER BY total DESC
        LIMIT 10
    """)

    mantenciones_pendientes = query_all("""
        SELECT * FROM mantenciones
        WHERE estado IN ('Pendiente','En proceso')
        ORDER BY fecha ASC, id DESC
        LIMIT 10
    """)

    recientes = query_all("""
        SELECT * FROM despachos
        ORDER BY id DESC
        LIMIT 10
    """)

    return render_page("Dashboard", r"""
    <div class="page-head">
        <div>
            <h1>Dashboard Admin</h1>
            <p>Resumen ejecutivo de despachos rápidos, mantenciones, usuarios y alertas operacionales.</p>
        </div>
    </div>

    <div class="grid grid-4">
        <div class="stat"><span>Total despachos</span><strong>{{ stats.total_despachos }}</strong></div>
        <div class="stat"><span>Entregados</span><strong>{{ stats.entregados }}</strong></div>
        <div class="stat"><span>Pendientes</span><strong>{{ stats.pendientes }}</strong></div>
        <div class="stat"><span>Monto documentado</span><strong>{{ stats.monto_total|money }}</strong></div>
        <div class="stat"><span>Mantenciones pendientes</span><strong>{{ stats.mant_pendientes }}</strong></div>
        <div class="stat"><span>Maquinarias con alerta</span><strong>{{ stats.maq_fuera }}</strong></div>
        <div class="stat"><span>Usuarios activos</span><strong>{{ stats.usuarios }}</strong></div>
    </div>

    <div class="grid grid-2" style="margin-top:18px;">
        <div class="card">
            <h3>Despachos por usuario</h3>
            <div class="table-wrap">
                <table>
                    <tr><th>Usuario</th><th>Total</th><th>Monto</th></tr>
                    {% for r in por_usuario %}
                    <tr>
                        <td>{{ r.usuario }}</td>
                        <td>{{ r.total }}</td>
                        <td>{{ r.monto|money }}</td>
                    </tr>
                    {% else %}
                    <tr><td colspan="3" class="muted">Sin datos.</td></tr>
                    {% endfor %}
                </table>
            </div>
        </div>

        <div class="card">
            <h3>Despachos por estado</h3>
            <div class="table-wrap">
                <table>
                    <tr><th>Estado</th><th>Total</th><th>Monto</th></tr>
                    {% for r in por_estado %}
                    <tr><td>{{ r.estado }}</td><td>{{ r.total }}</td><td>{{ r.monto|money }}</td></tr>
                    {% else %}
                    <tr><td colspan="3" class="muted">Sin datos.</td></tr>
                    {% endfor %}
                </table>
            </div>
        </div>
    </div>

    <div class="grid grid-2">
        <div class="card">
            <h3>Mantenciones pendientes</h3>
            <div class="table-wrap">
                <table>
                    <tr><th>Fecha</th><th>Maquinaria</th><th>Tipo</th><th>Estado</th></tr>
                    {% for m in mantenciones_pendientes %}
                    <tr>
                        <td>{{ m.fecha }}</td>
                        <td>{{ m.maquinaria_nombre }}</td>
                        <td>{{ m.tipo_mantencion }}</td>
                        <td><span class="badge warn">{{ m.estado }}</span></td>
                    </tr>
                    {% else %}
                    <tr><td colspan="4" class="muted">Sin mantenciones pendientes.</td></tr>
                    {% endfor %}
                </table>
            </div>
        </div>

        <div class="card">
            <h3>Últimos despachos</h3>
            <div class="table-wrap">
                <table>
                    <tr><th>Documento</th><th>Estado</th><th>Monto</th><th>Usuario</th><th>Fecha</th></tr>
                    {% for d in recientes %}
                    <tr>
                        <td>{{ d.tipo_documento }} {{ d.numero_documento }}</td>
                        <td><span class="badge {% if d.estado == 'Entregado' %}ok{% elif d.estado == 'Pendiente' %}warn{% else %}info{% endif %}">{{ d.estado }}</span></td>
                        <td>{{ d.monto|money }}</td>
                        <td>{{ d.usuario_nombre }}</td>
                        <td>{{ d.created_at }}</td>
                    </tr>
                    {% else %}
                    <tr><td colspan="5" class="muted">Sin registros.</td></tr>
                    {% endfor %}
                </table>
            </div>
        </div>
    </div>
    """, stats=stats, por_usuario=por_usuario, por_estado=por_estado,
       mantenciones_pendientes=mantenciones_pendientes, recientes=recientes)

# ============================================================
# DESPACHOS
# ============================================================

@app.route("/despachos", methods=["GET", "POST"])
@login_required
@permission_required("despachos")
def despachos():
    estados = ["Entregado", "Pendiente"]
    tipos = get_config_list("tipos_documento")

    if request.method == "POST":
        user = current_user()
        numero = request.form.get("numero_documento", "").strip()
        tipo = request.form.get("tipo_documento", "").strip()
        estado = request.form.get("estado", "Entregado").strip() or "Entregado"
        if estado not in estados:
            estado = "Entregado"

        if not numero or not tipo:
            flash("Número de documento y tipo de documento son obligatorios.", "error")
        else:
            new_id = insert_and_get_id("""
                INSERT INTO despachos (
                    numero_documento, tipo_documento, estado, cliente, telefono, destino,
                    patente, conductor, pioneta, monto, usuario_id, usuario_nombre,
                    created_at, updated_at
                )
                VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
            """, (
                numero,
                tipo,
                estado,
                "",
                "",
                "",
                "",
                "",
                "",
                money_to_float(request.form.get("monto")),
                user["id"],
                user["full_name"],
                now_str(),
                now_str(),
            ))
            write_audit("despachos", "crear", new_id, "registro", "", f"{tipo} {numero} / {estado}")
            flash("Despacho registrado correctamente.", "success")
            return redirect(url_for("despachos"))

    recientes = query_all("SELECT * FROM despachos ORDER BY id DESC LIMIT 25")
    audit = query_all("SELECT * FROM audit_log WHERE modulo='despachos' ORDER BY id DESC LIMIT 8") if is_admin() else []

    return render_page("Despachos", r"""
    <div class="page-head">
        <div>
            <h1>Módulo Despachos</h1>
            <p>Registro rápido para operadores: documento, tipo, estado y monto. El usuario queda registrado automáticamente.</p>
        </div>
    </div>

    <div class="card">
        <h2>Nuevo registro rápido</h2>
        <form method="post">
            <div class="form-row">
                <div>
                    <label>Número documento *</label>
                    <input name="numero_documento" required autofocus>
                </div>
                <div>
                    <label>Tipo documento *</label>
                    <select name="tipo_documento" required>
                        {% for t in tipos %}<option>{{ t }}</option>{% endfor %}
                    </select>
                </div>
                <div>
                    <label>Estado *</label>
                    <select name="estado" required>
                        <option value="Entregado" selected>Entregado</option>
                        <option value="Pendiente">Pendiente</option>
                    </select>
                </div>
                <div>
                    <label>Monto documento</label>
                    <input name="monto" inputmode="decimal" placeholder="Ej: 120000">
                </div>
            </div>

            <div class="actions">
                <button class="btn btn-primary">Guardar despacho</button>
                <a class="btn btn-secondary" href="{{ url_for('consulta') }}">Ir a consulta</a>
            </div>
        </form>
    </div>

    <div class="card">
        <div class="section-title">
            <h2>Registros recientes</h2>
            {% if is_admin %}<a class="btn btn-secondary btn-small" href="{{ url_for('export_despachos') }}">Exportar Excel</a>{% endif %}
        </div>
        <div class="table-wrap">
            <table>
                <tr>
                    <th>ID</th><th>Documento</th><th>Estado</th><th>Monto</th><th>Usuario</th><th>Fecha</th>{% if is_admin %}<th>Acción</th>{% endif %}
                </tr>
                {% for d in recientes %}
                <tr>
                    <td>{{ d.id }}</td>
                    <td>{{ d.tipo_documento }} {{ d.numero_documento }}</td>
                    <td><span class="badge {% if d.estado == 'Entregado' %}ok{% elif d.estado == 'Pendiente' %}warn{% else %}info{% endif %}">{{ d.estado }}</span></td>
                    <td>{{ d.monto|money }}</td>
                    <td>{{ d.usuario_nombre }}</td>
                    <td>{{ d.created_at }}</td>
                    {% if is_admin %}
                    <td><a class="btn btn-secondary btn-small" href="{{ url_for('edit_despacho', despacho_id=d.id) }}">Editar</a></td>
                    {% endif %}
                </tr>
                {% else %}
                <tr><td colspan="7" class="muted">Sin registros.</td></tr>
                {% endfor %}
            </table>
        </div>
    </div>

    {% if is_admin %}
    <div class="card">
        <h3>Auditoría reciente de despachos</h3>
        <div class="table-wrap">
            <table>
                <tr><th>Fecha</th><th>Usuario</th><th>Acción</th><th>Campo</th><th>Anterior</th><th>Nuevo</th></tr>
                {% for a in audit %}
                <tr>
                    <td>{{ a.created_at }}</td><td>{{ a.usuario_nombre }}</td><td>{{ a.accion }}</td>
                    <td>{{ a.campo }}</td><td>{{ a.valor_anterior }}</td><td>{{ a.valor_nuevo }}</td>
                </tr>
                {% else %}
                <tr><td colspan="6" class="muted">Sin auditoría.</td></tr>
                {% endfor %}
            </table>
        </div>
    </div>
    {% endif %}
    """, estados=estados, tipos=tipos, recientes=recientes, audit=audit, is_admin=is_admin())

@app.route("/despachos/<int:despacho_id>/editar", methods=["GET", "POST"])
@login_required
@permission_required("dashboard")
def edit_despacho(despacho_id):
    despacho = query_one("SELECT * FROM despachos WHERE id = ?", (despacho_id,))
    if not despacho:
        flash("Despacho no encontrado.", "error")
        return redirect(url_for("despachos"))

    estados = ["Entregado", "Pendiente"]
    tipos = get_config_list("tipos_documento")
    fields = ["numero_documento", "tipo_documento", "estado", "monto"]

    if request.method == "POST":
        estado = request.form.get("estado", "Entregado").strip() or "Entregado"
        if estado not in estados:
            estado = "Entregado"

        new_data = {
            "numero_documento": request.form.get("numero_documento", "").strip(),
            "tipo_documento": request.form.get("tipo_documento", "").strip(),
            "estado": estado,
            "monto": money_to_float(request.form.get("monto")),
        }

        for field in fields:
            old = despacho[field]
            new = new_data[field]
            if str(old or "") != str(new or ""):
                write_audit("despachos", "editar", despacho_id, field, old, new)

        execute("""
            UPDATE despachos
            SET numero_documento=?, tipo_documento=?, estado=?, monto=?, updated_at=?
            WHERE id=?
        """, (
            new_data["numero_documento"], new_data["tipo_documento"], new_data["estado"],
            new_data["monto"], now_str(), despacho_id
        ))
        flash("Despacho actualizado correctamente.", "success")
        return redirect(url_for("despachos"))

    return render_page("Editar despacho", r"""
    <div class="page-head">
        <div>
            <h1>Editar despacho #{{ despacho.id }}</h1>
            <p>Edición restringida a administradores. Cada cambio queda registrado en auditoría.</p>
        </div>
    </div>

    <div class="card">
        <form method="post">
            <div class="form-row">
                <div><label>Número documento</label><input name="numero_documento" value="{{ despacho.numero_documento }}" required></div>
                <div>
                    <label>Tipo documento</label>
                    <select name="tipo_documento">
                        {% for t in tipos %}<option {% if despacho.tipo_documento==t %}selected{% endif %}>{{ t }}</option>{% endfor %}
                    </select>
                </div>
                <div>
                    <label>Estado</label>
                    <select name="estado">
                        {% for e in estados %}<option {% if despacho.estado==e %}selected{% endif %}>{{ e }}</option>{% endfor %}
                    </select>
                </div>
                <div><label>Monto</label><input name="monto" value="{{ despacho.monto }}"></div>
            </div>
            <div class="actions">
                <button class="btn btn-primary">Guardar cambios</button>
                <a class="btn btn-secondary" href="{{ url_for('despachos') }}">Volver</a>
            </div>
        </form>
    </div>
    """, despacho=despacho, estados=estados, tipos=tipos)

# ============================================================
# CONSULTA
# ============================================================

@app.route("/consulta")
@login_required
@permission_required("consulta")
def consulta():
    q = request.args.get("q", "").strip()
    estado = request.args.get("estado", "").strip()
    desde = request.args.get("desde", "").strip()
    hasta = request.args.get("hasta", "").strip()

    wheres = []
    params = []
    if q:
        wheres.append("(numero_documento LIKE ? OR tipo_documento LIKE ? OR usuario_nombre LIKE ?)")
        params += [f"%{q}%", f"%{q}%", f"%{q}%"]
    if estado:
        wheres.append("estado = ?")
        params.append(estado)
    if desde:
        wheres.append("date(created_at) >= date(?)")
        params.append(desde)
    if hasta:
        wheres.append("date(created_at) <= date(?)")
        params.append(hasta)

    where_sql = "WHERE " + " AND ".join(wheres) if wheres else ""
    rows = query_all(f"""
        SELECT * FROM despachos
        {where_sql}
        ORDER BY id DESC
        LIMIT 500
    """, params)
    estados = ["Entregado", "Pendiente"]

    return render_page("Consulta", r"""
    <div class="page-head">
        <div>
            <h1>Consulta</h1>
            <p>Búsqueda de registros rápidos de despacho. Vista simple para operación.</p>
        </div>
    </div>

    <div class="card">
        <form method="get">
            <div class="form-row">
                <div>
                    <label>Buscar</label>
                    <input name="q" value="{{ request.args.get('q','') }}" placeholder="Documento, tipo o usuario">
                </div>
                <div>
                    <label>Estado</label>
                    <select name="estado">
                        <option value="">Todos</option>
                        {% for e in estados %}
                        <option value="{{ e }}" {% if request.args.get('estado')==e %}selected{% endif %}>{{ e }}</option>
                        {% endfor %}
                    </select>
                </div>
                <div>
                    <label>Fecha desde</label>
                    <input type="date" name="desde" value="{{ request.args.get('desde','') }}">
                </div>
                <div>
                    <label>Fecha hasta</label>
                    <input type="date" name="hasta" value="{{ request.args.get('hasta','') }}">
                </div>
            </div>
            <div class="actions">
                <button class="btn btn-primary">Filtrar</button>
                <a class="btn btn-secondary" href="{{ url_for('consulta') }}">Limpiar</a>
                {% if is_admin %}
                <a class="btn btn-secondary" href="{{ url_for('export_despachos', q=request.args.get('q',''), estado=request.args.get('estado',''), desde=request.args.get('desde',''), hasta=request.args.get('hasta','')) }}">Exportar resultado</a>
                {% endif %}
            </div>
        </form>
    </div>

    <div class="card">
        <h2>Resultado: {{ rows|length }} registros</h2>
        <div class="table-wrap">
            <table>
                <tr>
                    <th>ID</th><th>Documento</th><th>Estado</th><th>Monto</th><th>Usuario</th><th>Fecha</th>
                </tr>
                {% for d in rows %}
                <tr>
                    <td>{{ d.id }}</td>
                    <td>{{ d.tipo_documento }} {{ d.numero_documento }}</td>
                    <td><span class="badge {% if d.estado == 'Entregado' %}ok{% elif d.estado == 'Pendiente' %}warn{% else %}info{% endif %}">{{ d.estado }}</span></td>
                    <td>{{ d.monto|money }}</td>
                    <td>{{ d.usuario_nombre }}</td>
                    <td>{{ d.created_at }}</td>
                </tr>
                {% else %}
                <tr><td colspan="6" class="muted">Sin resultados.</td></tr>
                {% endfor %}
            </table>
        </div>
    </div>
    """, rows=rows, estados=estados, request=request, is_admin=is_admin())

# ============================================================
# MANTENCIONES
# ============================================================

@app.route("/mantenciones", methods=["GET", "POST"])
@login_required
@permission_required("mantenciones")
def mantenciones():
    maqs = query_all("SELECT * FROM maquinarias WHERE estado != 'Vendida / dada de baja' ORDER BY nombre")
    estados = get_config_list("estados_mantencion")
    tipos = get_config_list("tipos_mantencion")

    if request.method == "POST":
        user = current_user()
        maquinaria_id = request.form.get("maquinaria_id") or None
        maquinaria_nombre = request.form.get("maquinaria_nombre_manual", "").strip()

        if maquinaria_id:
            maq = query_one("SELECT * FROM maquinarias WHERE id = ?", (maquinaria_id,))
            maquinaria_nombre = maq["nombre"] if maq else maquinaria_nombre

        fecha = request.form.get("fecha", today_str())
        tipo_mantencion = request.form.get("tipo_mantencion", "").strip()
        estado = request.form.get("estado", "Pendiente").strip()

        if not maquinaria_nombre or not fecha or not tipo_mantencion:
            flash("Maquinaria, fecha y tipo de mantención son obligatorios.", "error")
        else:
            new_id = insert_and_get_id("""
                INSERT INTO mantenciones (
                    maquinaria_id, maquinaria_nombre, fecha, tipo_mantencion, estado,
                    responsable, costo, observaciones, usuario_id, usuario_nombre, created_at, updated_at
                )
                VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
            """, (
                maquinaria_id,
                maquinaria_nombre,
                fecha,
                tipo_mantencion,
                estado,
                request.form.get("responsable", "").strip(),
                money_to_float(request.form.get("costo")),
                request.form.get("observaciones", "").strip(),
                user["id"],
                user["full_name"],
                now_str(),
                now_str(),
            ))
            write_audit("mantenciones", "crear", new_id, "registro", "", maquinaria_nombre)
            flash("Mantención registrada correctamente.", "success")
            return redirect(url_for("mantenciones"))

    rows = query_all("SELECT * FROM mantenciones ORDER BY fecha DESC, id DESC LIMIT 100")

    return render_page("Mantenciones", r"""
    <div class="page-head">
        <div>
            <h1>Módulo Mantenciones</h1>
            <p>Control de mantenciones, maquinaria y alertas operativas.</p>
        </div>
    </div>

    <div class="card">
        <div class="section-title">
            <h2>Nueva mantención</h2>
            <a class="btn btn-secondary btn-small" href="{{ url_for('export_mantenciones') }}">Exportar Excel</a>
        </div>
        <form method="post">
            <div class="form-row">
                <div>
                    <label>Maquinaria existente</label>
                    <select name="maquinaria_id">
                        <option value="">Seleccionar o escribir manual</option>
                        {% for m in maqs %}
                        <option value="{{ m.id }}">{{ m.nombre }}{% if m.patente %} · {{ m.patente }}{% endif %}</option>
                        {% endfor %}
                    </select>
                </div>
                <div>
                    <label>Maquinaria manual</label>
                    <input name="maquinaria_nombre_manual" placeholder="Solo si no existe en listado">
                </div>
                <div>
                    <label>Fecha</label>
                    <input type="date" name="fecha" value="{{ today }}" required>
                </div>
                <div>
                    <label>Tipo</label>
                    <select name="tipo_mantencion" required>
                        {% for t in tipos %}<option>{{ t }}</option>{% endfor %}
                    </select>
                </div>
            </div>

            <div class="form-row" style="margin-top:14px;">
                <div>
                    <label>Estado</label>
                    <select name="estado">
                        {% for e in estados %}<option>{{ e }}</option>{% endfor %}
                    </select>
                </div>
                <div>
                    <label>Responsable</label>
                    <input name="responsable">
                </div>
                <div>
                    <label>Costo</label>
                    <input name="costo" inputmode="decimal">
                </div>
                <div>
                    <label>Usuario</label>
                    <input value="{{ user.full_name }}" disabled>
                </div>
            </div>

            <div style="margin-top:14px;">
                <label>Observaciones</label>
                <textarea name="observaciones"></textarea>
            </div>

            <div class="actions">
                <button class="btn btn-primary">Guardar mantención</button>
            </div>
        </form>
    </div>

    <div class="card">
        <h2>Historial de mantenciones</h2>
        <div class="table-wrap">
            <table>
                <tr>
                    <th>ID</th><th>Fecha</th><th>Maquinaria</th><th>Tipo</th><th>Estado</th>
                    <th>Responsable</th><th>Costo</th><th>Usuario</th><th>Obs.</th>{% if is_admin %}<th>Acción</th>{% endif %}
                </tr>
                {% for m in rows %}
                <tr>
                    <td>{{ m.id }}</td>
                    <td>{{ m.fecha }}</td>
                    <td>{{ m.maquinaria_nombre }}</td>
                    <td>{{ m.tipo_mantencion }}</td>
                    <td><span class="badge {% if m.estado == 'Realizada' %}ok{% elif m.estado in ['Pendiente','En proceso'] %}warn{% else %}info{% endif %}">{{ m.estado }}</span></td>
                    <td>{{ m.responsable }}</td>
                    <td>{{ m.costo|money }}</td>
                    <td>{{ m.usuario_nombre }}</td>
                    <td>{{ m.observaciones }}</td>
                    {% if is_admin %}<td><a class="btn btn-secondary btn-small" href="{{ url_for('edit_mantencion', mantencion_id=m.id) }}">Editar</a></td>{% endif %}
                </tr>
                {% else %}
                <tr><td colspan="10" class="muted">Sin mantenciones.</td></tr>
                {% endfor %}
            </table>
        </div>
    </div>
    """, maqs=maqs, estados=estados, tipos=tipos, rows=rows, today=today_str(), user=current_user(), is_admin=is_admin())


@app.route("/mantenciones/<int:mantencion_id>/editar", methods=["GET", "POST"])
@login_required
@permission_required("dashboard")
def edit_mantencion(mantencion_id):
    m = query_one("SELECT * FROM mantenciones WHERE id = ?", (mantencion_id,))
    if not m:
        flash("Mantención no encontrada.", "error")
        return redirect(url_for("mantenciones"))

    estados = get_config_list("estados_mantencion")
    tipos = get_config_list("tipos_mantencion")

    if request.method == "POST":
        data = {
            "maquinaria_nombre": request.form.get("maquinaria_nombre", "").strip(),
            "fecha": request.form.get("fecha", "").strip(),
            "tipo_mantencion": request.form.get("tipo_mantencion", "").strip(),
            "estado": request.form.get("estado", "").strip(),
            "responsable": request.form.get("responsable", "").strip(),
            "costo": money_to_float(request.form.get("costo")),
            "observaciones": request.form.get("observaciones", "").strip(),
        }
        for k, v in data.items():
            if str(m[k] or "") != str(v or ""):
                write_audit("mantenciones", "editar", mantencion_id, k, m[k], v)

        execute("""
            UPDATE mantenciones
            SET maquinaria_nombre=?, fecha=?, tipo_mantencion=?, estado=?, responsable=?, costo=?, observaciones=?, updated_at=?
            WHERE id=?
        """, (
            data["maquinaria_nombre"], data["fecha"], data["tipo_mantencion"], data["estado"],
            data["responsable"], data["costo"], data["observaciones"], now_str(), mantencion_id
        ))
        flash("Mantención actualizada.", "success")
        return redirect(url_for("mantenciones"))

    return render_page("Editar mantención", r"""
    <div class="page-head">
        <div>
            <h1>Editar mantención #{{ m.id }}</h1>
            <p>Edición restringida a administradores.</p>
        </div>
    </div>
    <div class="card">
        <form method="post">
            <div class="form-row">
                <div><label>Maquinaria</label><input name="maquinaria_nombre" value="{{ m.maquinaria_nombre }}" required></div>
                <div><label>Fecha</label><input type="date" name="fecha" value="{{ m.fecha }}" required></div>
                <div>
                    <label>Tipo</label>
                    <select name="tipo_mantencion">
                        {% for t in tipos %}<option {% if m.tipo_mantencion==t %}selected{% endif %}>{{ t }}</option>{% endfor %}
                    </select>
                </div>
                <div>
                    <label>Estado</label>
                    <select name="estado">
                        {% for e in estados %}<option {% if m.estado==e %}selected{% endif %}>{{ e }}</option>{% endfor %}
                    </select>
                </div>
            </div>
            <div class="form-row-2" style="margin-top:14px;">
                <div><label>Responsable</label><input name="responsable" value="{{ m.responsable }}"></div>
                <div><label>Costo</label><input name="costo" value="{{ m.costo }}"></div>
            </div>
            <div style="margin-top:14px;">
                <label>Observaciones</label>
                <textarea name="observaciones">{{ m.observaciones }}</textarea>
            </div>
            <div class="actions">
                <button class="btn btn-primary">Guardar cambios</button>
                <a class="btn btn-secondary" href="{{ url_for('mantenciones') }}">Volver</a>
            </div>
        </form>
    </div>
    """, m=m, estados=estados, tipos=tipos)


# ============================================================
# ADMINISTRACIÓN / CONFIGURACIÓN
# ============================================================

@app.route("/administracion", methods=["GET", "POST"])
@login_required
@permission_required("administracion")
def administracion():
    keys = [
        ("estados_despacho", "Estados de despacho"),
        ("tipos_documento", "Tipos de documento"),
        ("sucursales", "Sucursales"),
        ("estados_maquinaria", "Estados de maquinaria"),
        ("estados_mantencion", "Estados de mantención"),
        ("tipos_mantencion", "Tipos de mantención"),
    ]

    if request.method == "POST":
        for clave, _label in keys:
            raw = request.form.get(clave, "")
            values = [line.strip() for line in raw.splitlines() if line.strip()]
            old = get_config_list(clave)
            set_config_list(clave, values)
            write_audit("administracion", "configurar", None, clave, ", ".join(old), ", ".join(values))
        flash("Configuración actualizada.", "success")
        return redirect(url_for("administracion"))

    config = {clave: "\n".join(get_config_list(clave)) for clave, _ in keys}

    return render_page("Administración", r"""
    <div class="page-head">
        <div>
            <h1>Módulo Administración / Configuración</h1>
            <p>Centro de control del sistema. Visible solo para administradores.</p>
        </div>
    </div>

    <div class="grid grid-3">
        <div class="stat"><span>Usuarios y permisos</span><strong>{{ users_count }}</strong></div>
        <div class="stat"><span>Maquinarias</span><strong>{{ maqs_count }}</strong></div>
        <div class="stat"><span>Vehículos</span><strong>{{ vehs_count }}</strong></div>
    </div>

    <div class="card" style="margin-top:18px;">
        <h2>Configuración general</h2>
        <p class="muted">Escribe un valor por línea. Estos listados alimentan los formularios del sistema.</p>
        <form method="post">
            <div class="grid grid-2">
                {% for clave, label in keys %}
                <div>
                    <label>{{ label }}</label>
                    <textarea name="{{ clave }}">{{ config[clave] }}</textarea>
                </div>
                {% endfor %}
            </div>
            <div class="actions">
                <button class="btn btn-primary">Guardar configuración</button>
            </div>
        </form>
    </div>

    <div class="grid grid-3">
        <a class="card" href="{{ url_for('usuarios') }}">
            <h3>Usuarios y permisos</h3>
            <p class="muted">Crear usuarios, asignar roles y ocultar Dashboard a quienes no sean admin.</p>
        </a>
        <a class="card" href="{{ url_for('maquinarias') }}">
            <h3>Maquinarias</h3>
            <p class="muted">Crear, editar y desactivar maquinaria. Modificación solo admin.</p>
        </a>
        <a class="card" href="{{ url_for('vehiculos') }}">
            <h3>Vehículos / Patentes</h3>
            <p class="muted">Controlar patentes, vencimientos, revisión técnica y seguro.</p>
        </a>
    </div>
    """, keys=keys, config=config,
       users_count=query_one("SELECT COUNT(*) c FROM users")["c"],
       maqs_count=query_one("SELECT COUNT(*) c FROM maquinarias")["c"],
       vehs_count=query_one("SELECT COUNT(*) c FROM vehiculos")["c"])


# ============================================================
# USUARIOS
# ============================================================

@app.route("/usuarios", methods=["GET", "POST"])
@login_required
@permission_required("usuarios")
def usuarios():
    if request.method == "POST":
        username = request.form.get("username", "").strip()
        full_name = request.form.get("full_name", "").strip()
        password = request.form.get("password", "").strip()
        role = request.form.get("role", "operador")
        perms = request.form.getlist("permissions")

        if not username or not full_name or not password:
            flash("Usuario, nombre y contraseña son obligatorios.", "error")
        elif query_one("SELECT id FROM users WHERE username = ?", (username,)):
            flash("Ese usuario ya existe.", "error")
        else:
            permissions = {"all": True} if role == "admin" else {p: True for p in perms}
            new_id = insert_and_get_id("""
                INSERT INTO users (username, full_name, password_hash, role, is_active, permissions, created_at, updated_at)
                VALUES (?, ?, ?, ?, 1, ?, ?, ?)
            """, (
                username, full_name, generate_password_hash(password), role,
                json.dumps(permissions), now_str(), now_str()
            ))
            write_audit("usuarios", "crear", new_id, "usuario", "", username)
            flash("Usuario creado.", "success")
            return redirect(url_for("usuarios"))

    rows = query_all("SELECT * FROM users ORDER BY id")
    permission_labels = PERMISSION_LABELS

    return render_page("Usuarios", r"""
    <div class="page-head">
        <div>
            <h1>Usuarios y permisos</h1>
            <p>Dashboard, Auditoría y Administración quedan ocultos para usuarios no autorizados.</p>
        </div>
    </div>

    <div class="card">
        <h2>Crear usuario</h2>
        <form method="post">
            <div class="form-row">
                <div><label>Usuario</label><input name="username" required></div>
                <div><label>Nombre completo</label><input name="full_name" required></div>
                <div><label>Contraseña</label><input name="password" required></div>
                <div>
                    <label>Rol</label>
                    <select name="role">
                        <option value="operador">Operador</option>
                        <option value="admin">Administrador</option>
                    </select>
                </div>
            </div>

            <h3>Permisos del usuario operativo</h3>
            <div class="checkbox-grid">
                {% for key, label in permission_labels.items() %}
                    {% if key != "dashboard" and key != "auditoria" and key != "administracion" %}
                    <label class="check-item"><input type="checkbox" name="permissions" value="{{ key }}"> {{ label }}</label>
                    {% else %}
                    <label class="check-item"><input type="checkbox" name="permissions" value="{{ key }}"> {{ label }} <span class="muted">(admin recomendado)</span></label>
                    {% endif %}
                {% endfor %}
            </div>

            <div class="actions">
                <button class="btn btn-primary">Crear usuario</button>
            </div>
        </form>
    </div>

    <div class="card">
        <h2>Usuarios existentes</h2>
        <div class="table-wrap">
            <table>
                <tr><th>ID</th><th>Usuario</th><th>Nombre</th><th>Rol</th><th>Estado</th><th>Permisos</th><th>Acción</th></tr>
                {% for u in rows %}
                <tr>
                    <td>{{ u.id }}</td>
                    <td>{{ u.username }}</td>
                    <td>{{ u.full_name }}</td>
                    <td><span class="badge {% if u.role == 'admin' %}ok{% else %}info{% endif %}">{{ u.role }}</span></td>
                    <td>{% if u.is_active %}<span class="badge ok">Activo</span>{% else %}<span class="badge bad">Inactivo</span>{% endif %}</td>
                    <td class="muted">{{ u.permissions }}</td>
                    <td><a class="btn btn-secondary btn-small" href="{{ url_for('edit_usuario', user_id=u.id) }}">Editar</a></td>
                </tr>
                {% endfor %}
            </table>
        </div>
    </div>
    """, rows=rows, permission_labels=permission_labels)


@app.route("/usuarios/<int:user_id>/editar", methods=["GET", "POST"])
@login_required
@permission_required("usuarios")
def edit_usuario(user_id):
    u = query_one("SELECT * FROM users WHERE id = ?", (user_id,))
    if not u:
        flash("Usuario no encontrado.", "error")
        return redirect(url_for("usuarios"))

    current_perms = user_permissions(u)

    if request.method == "POST":
        full_name = request.form.get("full_name", "").strip()
        role = request.form.get("role", "operador")
        is_active = 1 if request.form.get("is_active") == "1" else 0
        password = request.form.get("password", "").strip()
        perms = request.form.getlist("permissions")
        permissions = {"all": True} if role == "admin" else {p: True for p in perms}

        # Evita que el usuario actual se desactive a sí mismo por accidente.
        if u["id"] == current_user()["id"] and not is_active:
            flash("No puedes desactivar tu propio usuario desde esta pantalla.", "error")
            return redirect(url_for("edit_usuario", user_id=user_id))

        if password:
            execute("""
                UPDATE users
                SET full_name=?, role=?, is_active=?, permissions=?, password_hash=?, updated_at=?
                WHERE id=?
            """, (
                full_name, role, is_active, json.dumps(permissions),
                generate_password_hash(password), now_str(), user_id
            ))
            write_audit("usuarios", "editar", user_id, "password", "anterior", "actualizada")
        else:
            execute("""
                UPDATE users
                SET full_name=?, role=?, is_active=?, permissions=?, updated_at=?
                WHERE id=?
            """, (
                full_name, role, is_active, json.dumps(permissions), now_str(), user_id
            ))

        write_audit("usuarios", "editar", user_id, "usuario", u["username"], f"{full_name} / {role}")
        flash("Usuario actualizado.", "success")
        return redirect(url_for("usuarios"))

    return render_page("Editar usuario", r"""
    <div class="page-head">
        <div>
            <h1>Editar usuario: {{ u.username }}</h1>
            <p>Los permisos controlan qué botones y secciones ve cada persona.</p>
        </div>
    </div>

    <div class="card">
        <form method="post">
            <div class="form-row">
                <div><label>Usuario</label><input value="{{ u.username }}" disabled></div>
                <div><label>Nombre completo</label><input name="full_name" value="{{ u.full_name }}" required></div>
                <div><label>Nueva contraseña</label><input name="password" placeholder="Dejar vacío para mantener"></div>
                <div>
                    <label>Rol</label>
                    <select name="role">
                        <option value="operador" {% if u.role=='operador' %}selected{% endif %}>Operador</option>
                        <option value="admin" {% if u.role=='admin' %}selected{% endif %}>Administrador</option>
                    </select>
                </div>
            </div>

            <div style="margin-top:14px;">
                <label>Estado</label>
                <select name="is_active">
                    <option value="1" {% if u.is_active %}selected{% endif %}>Activo</option>
                    <option value="0" {% if not u.is_active %}selected{% endif %}>Inactivo</option>
                </select>
            </div>

            <h3>Permisos</h3>
            <div class="checkbox-grid">
                {% for key, label in permission_labels.items() %}
                <label class="check-item">
                    <input type="checkbox" name="permissions" value="{{ key }}" {% if current_perms.get(key) or current_perms.get('all') %}checked{% endif %}>
                    {{ label }}
                </label>
                {% endfor %}
            </div>

            <div class="actions">
                <button class="btn btn-primary">Guardar usuario</button>
                <a class="btn btn-secondary" href="{{ url_for('usuarios') }}">Volver</a>
            </div>
        </form>
    </div>
    """, u=u, current_perms=current_perms, permission_labels=PERMISSION_LABELS)


# ============================================================
# MAQUINARIAS
# ============================================================

@app.route("/maquinarias", methods=["GET", "POST"])
@login_required
@permission_required("maquinarias")
def maquinarias():
    estados = get_config_list("estados_maquinaria")

    if request.method == "POST":
        new_id = insert_and_get_id("""
            INSERT INTO maquinarias (codigo, nombre, tipo, marca, modelo, anio, patente, estado, observaciones, created_at, updated_at)
            VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
        """, (
            request.form.get("codigo", "").strip(),
            request.form.get("nombre", "").strip(),
            request.form.get("tipo", "").strip(),
            request.form.get("marca", "").strip(),
            request.form.get("modelo", "").strip(),
            request.form.get("anio", "").strip(),
            request.form.get("patente", "").strip(),
            request.form.get("estado", "Activa").strip(),
            request.form.get("observaciones", "").strip(),
            now_str(),
            now_str(),
        ))
        write_audit("maquinarias", "crear", new_id, "maquinaria", "", request.form.get("nombre", ""))
        flash("Maquinaria creada.", "success")
        return redirect(url_for("maquinarias"))

    rows = query_all("SELECT * FROM maquinarias ORDER BY estado, nombre")

    return render_page("Maquinarias", r"""
    <div class="page-head">
        <div>
            <h1>Maquinarias</h1>
            <p>Crear, editar o desactivar maquinaria. Esta sección es solo para administradores.</p>
        </div>
    </div>

    <div class="card">
        <h2>Nueva maquinaria</h2>
        <form method="post">
            <div class="form-row">
                <div><label>Código interno</label><input name="codigo"></div>
                <div><label>Nombre *</label><input name="nombre" required></div>
                <div><label>Tipo</label><input name="tipo" placeholder="Grúa, camión, elevador, etc."></div>
                <div><label>Estado</label>
                    <select name="estado">{% for e in estados %}<option>{{ e }}</option>{% endfor %}</select>
                </div>
            </div>
            <div class="form-row" style="margin-top:14px;">
                <div><label>Marca</label><input name="marca"></div>
                <div><label>Modelo</label><input name="modelo"></div>
                <div><label>Año</label><input name="anio"></div>
                <div><label>Patente</label><input name="patente"></div>
            </div>
            <div style="margin-top:14px;">
                <label>Observaciones</label>
                <textarea name="observaciones"></textarea>
            </div>
            <div class="actions"><button class="btn btn-primary">Guardar maquinaria</button></div>
        </form>
    </div>

    <div class="card">
        <div class="section-title">
            <h2>Listado de maquinarias</h2>
            <a class="btn btn-secondary btn-small" href="{{ url_for('export_maquinarias') }}">Exportar Excel</a>
        </div>
        <div class="table-wrap">
            <table>
                <tr><th>Código</th><th>Nombre</th><th>Tipo</th><th>Marca/Modelo</th><th>Patente</th><th>Estado</th><th>Obs.</th><th>Acción</th></tr>
                {% for m in rows %}
                <tr>
                    <td>{{ m.codigo }}</td>
                    <td>{{ m.nombre }}</td>
                    <td>{{ m.tipo }}</td>
                    <td>{{ m.marca }} {{ m.modelo }} {{ m.anio }}</td>
                    <td>{{ m.patente }}</td>
                    <td><span class="badge {% if m.estado == 'Activa' %}ok{% elif m.estado == 'Fuera de servicio' %}bad{% else %}warn{% endif %}">{{ m.estado }}</span></td>
                    <td>{{ m.observaciones }}</td>
                    <td><a class="btn btn-secondary btn-small" href="{{ url_for('edit_maquinaria', maquinaria_id=m.id) }}">Editar</a></td>
                </tr>
                {% else %}
                <tr><td colspan="8" class="muted">Sin maquinarias.</td></tr>
                {% endfor %}
            </table>
        </div>
    </div>
    """, rows=rows, estados=estados)


@app.route("/maquinarias/<int:maquinaria_id>/editar", methods=["GET", "POST"])
@login_required
@permission_required("maquinarias")
def edit_maquinaria(maquinaria_id):
    m = query_one("SELECT * FROM maquinarias WHERE id = ?", (maquinaria_id,))
    if not m:
        flash("Maquinaria no encontrada.", "error")
        return redirect(url_for("maquinarias"))
    estados = get_config_list("estados_maquinaria")

    if request.method == "POST":
        data = {
            "codigo": request.form.get("codigo", "").strip(),
            "nombre": request.form.get("nombre", "").strip(),
            "tipo": request.form.get("tipo", "").strip(),
            "marca": request.form.get("marca", "").strip(),
            "modelo": request.form.get("modelo", "").strip(),
            "anio": request.form.get("anio", "").strip(),
            "patente": request.form.get("patente", "").strip(),
            "estado": request.form.get("estado", "").strip(),
            "observaciones": request.form.get("observaciones", "").strip(),
        }
        for k, v in data.items():
            if str(m[k] or "") != str(v or ""):
                write_audit("maquinarias", "editar", maquinaria_id, k, m[k], v)
        execute("""
            UPDATE maquinarias
            SET codigo=?, nombre=?, tipo=?, marca=?, modelo=?, anio=?, patente=?, estado=?, observaciones=?, updated_at=?
            WHERE id=?
        """, (
            data["codigo"], data["nombre"], data["tipo"], data["marca"], data["modelo"], data["anio"],
            data["patente"], data["estado"], data["observaciones"], now_str(), maquinaria_id
        ))
        flash("Maquinaria actualizada.", "success")
        return redirect(url_for("maquinarias"))

    return render_page("Editar maquinaria", r"""
    <div class="page-head"><div><h1>Editar maquinaria</h1><p>Modificación solo admin.</p></div></div>
    <div class="card">
        <form method="post">
            <div class="form-row">
                <div><label>Código</label><input name="codigo" value="{{ m.codigo }}"></div>
                <div><label>Nombre</label><input name="nombre" value="{{ m.nombre }}" required></div>
                <div><label>Tipo</label><input name="tipo" value="{{ m.tipo }}"></div>
                <div><label>Estado</label><select name="estado">{% for e in estados %}<option {% if m.estado==e %}selected{% endif %}>{{ e }}</option>{% endfor %}</select></div>
            </div>
            <div class="form-row" style="margin-top:14px;">
                <div><label>Marca</label><input name="marca" value="{{ m.marca }}"></div>
                <div><label>Modelo</label><input name="modelo" value="{{ m.modelo }}"></div>
                <div><label>Año</label><input name="anio" value="{{ m.anio }}"></div>
                <div><label>Patente</label><input name="patente" value="{{ m.patente }}"></div>
            </div>
            <div style="margin-top:14px;"><label>Observaciones</label><textarea name="observaciones">{{ m.observaciones }}</textarea></div>
            <div class="actions">
                <button class="btn btn-primary">Guardar cambios</button>
                <a class="btn btn-secondary" href="{{ url_for('maquinarias') }}">Volver</a>
            </div>
        </form>
    </div>
    """, m=m, estados=estados)


# ============================================================
# VEHÍCULOS / PATENTES
# ============================================================

@app.route("/vehiculos", methods=["GET", "POST"])
@login_required
@permission_required("vehiculos")
def vehiculos():
    if request.method == "POST":
        patente = request.form.get("patente", "").strip().upper()
        if not patente:
            flash("La patente es obligatoria.", "error")
        elif query_one("SELECT id FROM vehiculos WHERE patente = ?", (patente,)):
            flash("Esa patente ya existe.", "error")
        else:
            new_id = insert_and_get_id("""
                INSERT INTO vehiculos (
                    patente, tipo, marca, modelo, anio, estado,
                    permiso_circulacion_vencimiento, revision_tecnica_vencimiento,
                    seguro_obligatorio_vencimiento, observaciones, created_at, updated_at
                )
                VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
            """, (
                patente,
                request.form.get("tipo", "").strip(),
                request.form.get("marca", "").strip(),
                request.form.get("modelo", "").strip(),
                request.form.get("anio", "").strip(),
                request.form.get("estado", "Activo").strip(),
                request.form.get("permiso_circulacion_vencimiento", "").strip(),
                request.form.get("revision_tecnica_vencimiento", "").strip(),
                request.form.get("seguro_obligatorio_vencimiento", "").strip(),
                request.form.get("observaciones", "").strip(),
                now_str(),
                now_str(),
            ))
            write_audit("vehiculos", "crear", new_id, "patente", "", patente)
            flash("Vehículo creado.", "success")
            return redirect(url_for("vehiculos"))

    rows = query_all("SELECT * FROM vehiculos ORDER BY patente")

    return render_page("Vehículos", r"""
    <div class="page-head"><div><h1>Vehículos / Patentes</h1><p>Control de vehículos, documentos y vencimientos.</p></div></div>

    <div class="card">
        <h2>Nuevo vehículo</h2>
        <form method="post">
            <div class="form-row">
                <div><label>Patente *</label><input name="patente" required></div>
                <div><label>Tipo</label><input name="tipo" placeholder="Camión, camioneta, auto"></div>
                <div><label>Marca</label><input name="marca"></div>
                <div><label>Modelo</label><input name="modelo"></div>
            </div>
            <div class="form-row" style="margin-top:14px;">
                <div><label>Año</label><input name="anio"></div>
                <div><label>Estado</label><select name="estado"><option>Activo</option><option>En mantención</option><option>Fuera de servicio</option><option>Vendido</option></select></div>
                <div><label>Permiso circulación</label><input type="date" name="permiso_circulacion_vencimiento"></div>
                <div><label>Revisión técnica</label><input type="date" name="revision_tecnica_vencimiento"></div>
            </div>
            <div class="form-row-2" style="margin-top:14px;">
                <div><label>Seguro obligatorio</label><input type="date" name="seguro_obligatorio_vencimiento"></div>
                <div><label>Observaciones</label><input name="observaciones"></div>
            </div>
            <div class="actions"><button class="btn btn-primary">Guardar vehículo</button></div>
        </form>
    </div>

    <div class="card">
        <div class="section-title">
            <h2>Listado de vehículos</h2>
            <a class="btn btn-secondary btn-small" href="{{ url_for('export_vehiculos') }}">Exportar Excel</a>
        </div>
        <div class="table-wrap">
            <table>
                <tr><th>Patente</th><th>Tipo</th><th>Marca/Modelo</th><th>Estado</th><th>Permiso</th><th>Rev. técnica</th><th>SOAP</th><th>Acción</th></tr>
                {% for v in rows %}
                <tr>
                    <td>{{ v.patente }}</td>
                    <td>{{ v.tipo }}</td>
                    <td>{{ v.marca }} {{ v.modelo }} {{ v.anio }}</td>
                    <td><span class="badge {% if v.estado == 'Activo' %}ok{% elif v.estado == 'Fuera de servicio' %}bad{% else %}warn{% endif %}">{{ v.estado }}</span></td>
                    <td>{{ v.permiso_circulacion_vencimiento }}</td>
                    <td>{{ v.revision_tecnica_vencimiento }}</td>
                    <td>{{ v.seguro_obligatorio_vencimiento }}</td>
                    <td><a class="btn btn-secondary btn-small" href="{{ url_for('edit_vehiculo', vehiculo_id=v.id) }}">Editar</a></td>
                </tr>
                {% else %}
                <tr><td colspan="8" class="muted">Sin vehículos.</td></tr>
                {% endfor %}
            </table>
        </div>
    </div>
    """, rows=rows)


@app.route("/vehiculos/<int:vehiculo_id>/editar", methods=["GET", "POST"])
@login_required
@permission_required("vehiculos")
def edit_vehiculo(vehiculo_id):
    v = query_one("SELECT * FROM vehiculos WHERE id = ?", (vehiculo_id,))
    if not v:
        flash("Vehículo no encontrado.", "error")
        return redirect(url_for("vehiculos"))

    if request.method == "POST":
        data = {
            "patente": request.form.get("patente", "").strip().upper(),
            "tipo": request.form.get("tipo", "").strip(),
            "marca": request.form.get("marca", "").strip(),
            "modelo": request.form.get("modelo", "").strip(),
            "anio": request.form.get("anio", "").strip(),
            "estado": request.form.get("estado", "").strip(),
            "permiso_circulacion_vencimiento": request.form.get("permiso_circulacion_vencimiento", "").strip(),
            "revision_tecnica_vencimiento": request.form.get("revision_tecnica_vencimiento", "").strip(),
            "seguro_obligatorio_vencimiento": request.form.get("seguro_obligatorio_vencimiento", "").strip(),
            "observaciones": request.form.get("observaciones", "").strip(),
        }
        for k, value in data.items():
            if str(v[k] or "") != str(value or ""):
                write_audit("vehiculos", "editar", vehiculo_id, k, v[k], value)
        execute("""
            UPDATE vehiculos
            SET patente=?, tipo=?, marca=?, modelo=?, anio=?, estado=?,
                permiso_circulacion_vencimiento=?, revision_tecnica_vencimiento=?,
                seguro_obligatorio_vencimiento=?, observaciones=?, updated_at=?
            WHERE id=?
        """, (
            data["patente"], data["tipo"], data["marca"], data["modelo"], data["anio"], data["estado"],
            data["permiso_circulacion_vencimiento"], data["revision_tecnica_vencimiento"],
            data["seguro_obligatorio_vencimiento"], data["observaciones"], now_str(), vehiculo_id
        ))
        flash("Vehículo actualizado.", "success")
        return redirect(url_for("vehiculos"))

    return render_page("Editar vehículo", r"""
    <div class="page-head"><div><h1>Editar vehículo</h1><p>Control administrativo de patente y vencimientos.</p></div></div>
    <div class="card">
        <form method="post">
            <div class="form-row">
                <div><label>Patente</label><input name="patente" value="{{ v.patente }}" required></div>
                <div><label>Tipo</label><input name="tipo" value="{{ v.tipo }}"></div>
                <div><label>Marca</label><input name="marca" value="{{ v.marca }}"></div>
                <div><label>Modelo</label><input name="modelo" value="{{ v.modelo }}"></div>
            </div>
            <div class="form-row" style="margin-top:14px;">
                <div><label>Año</label><input name="anio" value="{{ v.anio }}"></div>
                <div><label>Estado</label><select name="estado">
                    {% for e in ['Activo','En mantención','Fuera de servicio','Vendido'] %}
                    <option {% if v.estado==e %}selected{% endif %}>{{ e }}</option>
                    {% endfor %}
                </select></div>
                <div><label>Permiso circulación</label><input type="date" name="permiso_circulacion_vencimiento" value="{{ v.permiso_circulacion_vencimiento }}"></div>
                <div><label>Revisión técnica</label><input type="date" name="revision_tecnica_vencimiento" value="{{ v.revision_tecnica_vencimiento }}"></div>
            </div>
            <div class="form-row-2" style="margin-top:14px;">
                <div><label>Seguro obligatorio</label><input type="date" name="seguro_obligatorio_vencimiento" value="{{ v.seguro_obligatorio_vencimiento }}"></div>
                <div><label>Observaciones</label><input name="observaciones" value="{{ v.observaciones }}"></div>
            </div>
            <div class="actions">
                <button class="btn btn-primary">Guardar cambios</button>
                <a class="btn btn-secondary" href="{{ url_for('vehiculos') }}">Volver</a>
            </div>
        </form>
    </div>
    """, v=v)


# ============================================================
# CONDUCTORES / PIONETAS
# ============================================================

@app.route("/logistica", methods=["GET", "POST"])
@login_required
@permission_required("logistica")
def logistica():
    if request.method == "POST":
        new_id = insert_and_get_id("""
            INSERT INTO personas_logistica (nombre, tipo, estado, telefono, observaciones, created_at, updated_at)
            VALUES (?, ?, ?, ?, ?, ?, ?)
        """, (
            request.form.get("nombre", "").strip(),
            request.form.get("tipo", "Conductor").strip(),
            request.form.get("estado", "Activo").strip(),
            request.form.get("telefono", "").strip(),
            request.form.get("observaciones", "").strip(),
            now_str(),
            now_str(),
        ))
        write_audit("logistica", "crear", new_id, "persona", "", request.form.get("nombre", ""))
        flash("Persona logística creada.", "success")
        return redirect(url_for("logistica"))

    rows = query_all("SELECT * FROM personas_logistica ORDER BY tipo, nombre")
    return render_page("Conductores y pionetas", r"""
    <div class="page-head"><div><h1>Conductores y pionetas</h1><p>Mantiene limpio el listado de personas para despachos internos.</p></div></div>

    <div class="card">
        <h2>Nueva persona</h2>
        <form method="post">
            <div class="form-row">
                <div><label>Nombre</label><input name="nombre" required></div>
                <div><label>Tipo</label><select name="tipo"><option>Conductor</option><option>Pioneta</option></select></div>
                <div><label>Estado</label><select name="estado"><option>Activo</option><option>Inactivo</option></select></div>
                <div><label>Teléfono</label><input name="telefono"></div>
            </div>
            <div style="margin-top:14px;"><label>Observaciones</label><input name="observaciones"></div>
            <div class="actions"><button class="btn btn-primary">Guardar</button></div>
        </form>
    </div>

    <div class="card">
        <h2>Listado</h2>
        <div class="table-wrap">
            <table>
                <tr><th>Nombre</th><th>Tipo</th><th>Estado</th><th>Teléfono</th><th>Obs.</th><th>Acción</th></tr>
                {% for p in rows %}
                <tr>
                    <td>{{ p.nombre }}</td>
                    <td>{{ p.tipo }}</td>
                    <td><span class="badge {% if p.estado == 'Activo' %}ok{% else %}bad{% endif %}">{{ p.estado }}</span></td>
                    <td>{{ p.telefono }}</td>
                    <td>{{ p.observaciones }}</td>
                    <td><a class="btn btn-secondary btn-small" href="{{ url_for('edit_logistica', persona_id=p.id) }}">Editar</a></td>
                </tr>
                {% else %}
                <tr><td colspan="6" class="muted">Sin personas.</td></tr>
                {% endfor %}
            </table>
        </div>
    </div>
    """, rows=rows)


@app.route("/logistica/<int:persona_id>/editar", methods=["GET", "POST"])
@login_required
@permission_required("logistica")
def edit_logistica(persona_id):
    p = query_one("SELECT * FROM personas_logistica WHERE id = ?", (persona_id,))
    if not p:
        flash("Persona no encontrada.", "error")
        return redirect(url_for("logistica"))

    if request.method == "POST":
        data = {
            "nombre": request.form.get("nombre", "").strip(),
            "tipo": request.form.get("tipo", "").strip(),
            "estado": request.form.get("estado", "").strip(),
            "telefono": request.form.get("telefono", "").strip(),
            "observaciones": request.form.get("observaciones", "").strip(),
        }
        for k, value in data.items():
            if str(p[k] or "") != str(value or ""):
                write_audit("logistica", "editar", persona_id, k, p[k], value)
        execute("""
            UPDATE personas_logistica
            SET nombre=?, tipo=?, estado=?, telefono=?, observaciones=?, updated_at=?
            WHERE id=?
        """, (data["nombre"], data["tipo"], data["estado"], data["telefono"], data["observaciones"], now_str(), persona_id))
        flash("Persona actualizada.", "success")
        return redirect(url_for("logistica"))

    return render_page("Editar persona logística", r"""
    <div class="page-head"><div><h1>Editar conductor / pioneta</h1></div></div>
    <div class="card">
        <form method="post">
            <div class="form-row">
                <div><label>Nombre</label><input name="nombre" value="{{ p.nombre }}" required></div>
                <div><label>Tipo</label><select name="tipo"><option {% if p.tipo=='Conductor' %}selected{% endif %}>Conductor</option><option {% if p.tipo=='Pioneta' %}selected{% endif %}>Pioneta</option></select></div>
                <div><label>Estado</label><select name="estado"><option {% if p.estado=='Activo' %}selected{% endif %}>Activo</option><option {% if p.estado=='Inactivo' %}selected{% endif %}>Inactivo</option></select></div>
                <div><label>Teléfono</label><input name="telefono" value="{{ p.telefono }}"></div>
            </div>
            <div style="margin-top:14px;"><label>Observaciones</label><input name="observaciones" value="{{ p.observaciones }}"></div>
            <div class="actions">
                <button class="btn btn-primary">Guardar cambios</button>
                <a class="btn btn-secondary" href="{{ url_for('logistica') }}">Volver</a>
            </div>
        </form>
    </div>
    """, p=p)


# ============================================================
# AUDITORÍA
# ============================================================

@app.route("/auditoria")
@login_required
@permission_required("auditoria")
def auditoria():
    modulo = request.args.get("modulo", "").strip()
    usuario = request.args.get("usuario", "").strip()

    wheres = []
    params = []
    if modulo:
        wheres.append("modulo = ?")
        params.append(modulo)
    if usuario:
        wheres.append("usuario_nombre LIKE ?")
        params.append(f"%{usuario}%")

    where_sql = "WHERE " + " AND ".join(wheres) if wheres else ""
    rows = query_all(f"""
        SELECT * FROM audit_log
        {where_sql}
        ORDER BY id DESC
        LIMIT 1000
    """, params)
    modulos = query_all("SELECT DISTINCT modulo FROM audit_log ORDER BY modulo")

    return render_page("Auditoría", r"""
    <div class="page-head">
        <div>
            <h1>Auditoría</h1>
            <p>Registro de creación, edición, configuración y accesos relevantes del sistema.</p>
        </div>
    </div>

    <div class="card">
        <form method="get">
            <div class="form-row-3">
                <div>
                    <label>Módulo</label>
                    <select name="modulo">
                        <option value="">Todos</option>
                        {% for m in modulos %}
                        <option value="{{ m.modulo }}" {% if request.args.get('modulo')==m.modulo %}selected{% endif %}>{{ m.modulo }}</option>
                        {% endfor %}
                    </select>
                </div>
                <div><label>Usuario</label><input name="usuario" value="{{ request.args.get('usuario','') }}"></div>
                <div class="actions" style="margin-top:24px;">
                    <button class="btn btn-primary">Filtrar</button>
                    <a class="btn btn-secondary" href="{{ url_for('auditoria') }}">Limpiar</a>
                    <a class="btn btn-secondary" href="{{ url_for('export_auditoria') }}">Exportar Excel</a>
                </div>
            </div>
        </form>
    </div>

    <div class="card">
        <h2>Eventos: {{ rows|length }}</h2>
        <div class="table-wrap">
            <table>
                <tr><th>ID</th><th>Fecha</th><th>Módulo</th><th>Acción</th><th>Registro</th><th>Campo</th><th>Anterior</th><th>Nuevo</th><th>Usuario</th></tr>
                {% for a in rows %}
                <tr>
                    <td>{{ a.id }}</td>
                    <td>{{ a.created_at }}</td>
                    <td>{{ a.modulo }}</td>
                    <td>{{ a.accion }}</td>
                    <td>{{ a.registro_id }}</td>
                    <td>{{ a.campo }}</td>
                    <td>{{ a.valor_anterior }}</td>
                    <td>{{ a.valor_nuevo }}</td>
                    <td>{{ a.usuario_nombre }}</td>
                </tr>
                {% else %}
                <tr><td colspan="9" class="muted">Sin eventos.</td></tr>
                {% endfor %}
            </table>
        </div>
    </div>
    """, rows=rows, modulos=modulos, request=request)



# ============================================================
# PRODUCTOS / MAESTRA DE PRECIOS
# ============================================================

@app.route("/productos", methods=["GET", "POST"])
@login_required
@permission_required("productos")
def productos():
    if request.method == "POST":
        file = request.files.get("archivo_productos")
        if not file or not file.filename:
            flash("Debes seleccionar un archivo Excel.", "error")
            return redirect(url_for("productos"))

        if not file.filename.lower().endswith((".xlsx", ".xlsm")):
            flash("Formato no permitido. Sube un archivo .xlsx o .xlsm.", "error")
            return redirect(url_for("productos"))

        try:
            resultado = procesar_excel_productos(file)
            flash(
                f"Productos actualizados: {resultado['total']} filas, {resultado['creados']} creados, "
                f"{resultado['actualizados']} actualizados, {resultado['errores']} errores.",
                "success"
            )
            return redirect(url_for("productos"))
        except Exception as exc:
            flash(f"Error al importar productos: {exc}", "error")
            return redirect(url_for("productos"))

    q = request.args.get("q", "").strip()
    wheres = []
    params = []
    if q:
        q_norm = normalize_text(q)
        tokens = tokenize_search(q)
        if tokens:
            wheres.append("(" + " OR ".join(["descripcion_busqueda LIKE ?" for _ in tokens[:5]]) + " OR codigo_producto LIKE ?)")
            params.extend([f"%{t}%" for t in tokens[:5]])
            params.append(f"%{q}%")
        else:
            wheres.append("codigo_producto LIKE ?")
            params.append(f"%{q}%")

    where_sql = "WHERE " + " AND ".join(wheres) if wheres else ""
    rows = query_all(f"""
        SELECT * FROM productos
        {where_sql}
        ORDER BY activo DESC, updated_at DESC
        LIMIT 250
    """, params)

    stats = {
        "total": query_one("SELECT COUNT(*) c FROM productos")["c"],
        "activos": query_one("SELECT COUNT(*) c FROM productos WHERE activo = 1")["c"],
        "con_stock": query_one("SELECT COUNT(*) c FROM productos WHERE activo = 1 AND stock > 0")["c"],
    }
    last_import = query_one("SELECT * FROM producto_importaciones ORDER BY id DESC LIMIT 1")
    history = query_all("SELECT * FROM producto_importaciones ORDER BY id DESC LIMIT 10")

    return render_page("Productos", r"""
    <div class="page-head">
        <div>
            <h1>Productos / Maestra de precios</h1>
            <p>Importa la planilla maestra para que Elias pueda cotizar con precio, stock, margen y contribución.</p>
        </div>
    </div>

    <div class="grid grid-3">
        <div class="stat"><span>Total productos</span><strong>{{ stats.total }}</strong></div>
        <div class="stat"><span>Activos</span><strong>{{ stats.activos }}</strong></div>
        <div class="stat"><span>Activos con stock</span><strong>{{ stats.con_stock }}</strong></div>
    </div>

    <div class="card" style="margin-top:18px;">
        <div class="section-title">
            <div>
                <h2>Actualizar productos desde Excel</h2>
                {% if last_import %}
                <p class="muted">Última actualización: <b>{{ last_import.created_at }}</b> por {{ last_import.usuario_nombre }} · {{ last_import.archivo_nombre }}</p>
                {% else %}
                <p class="muted">Aún no existe una importación de productos.</p>
                {% endif %}
            </div>
            <a class="btn btn-secondary btn-small" href="{{ url_for('plantilla_productos') }}">Descargar plantilla</a>
        </div>

        <form method="post" enctype="multipart/form-data">
            <div class="form-row-2">
                <div>
                    <label>Archivo Excel de productos</label>
                    <input type="file" name="archivo_productos" accept=".xlsx,.xlsm" required>
                </div>
                <div class="placeholder">
                    Columnas mínimas esperadas:<br>
                    <b>Código Producto</b>, <b>Descripción</b>, <b>Precio Compra Neto</b>, <b>Precio Venta Bruto</b>, <b>Stock</b>, <b>Activo</b>.
                </div>
            </div>
            <div class="actions">
                <button class="btn btn-primary">Importar / actualizar productos</button>
                <a class="btn btn-secondary" href="{{ url_for('export_productos') }}">Exportar productos</a>
            </div>
        </form>
    </div>

    <div class="card">
        <h2>Buscar productos</h2>
        <form method="get">
            <div class="form-row-2">
                <div><label>Código o descripción</label><input name="q" value="{{ request.args.get('q','') }}" placeholder="Ej: cemento, 123456, plancha OSB"></div>
                <div class="actions" style="margin-top:24px;">
                    <button class="btn btn-primary">Buscar</button>
                    <a class="btn btn-secondary" href="{{ url_for('productos') }}">Limpiar</a>
                </div>
            </div>
        </form>
        <div class="table-wrap" style="margin-top:14px;">
            <table>
                <tr><th>Código</th><th>Descripción</th><th>Compra neto</th><th>Venta bruto</th><th>Stock</th><th>Activo</th><th>Actualizado</th></tr>
                {% for p in rows %}
                <tr>
                    <td>{{ p.codigo_producto }}</td>
                    <td>{{ p.descripcion }}</td>
                    <td>{{ p.precio_compra_neto|money }}</td>
                    <td>{{ p.precio_venta_bruto|money }}</td>
                    <td>{{ "%.2f"|format(p.stock or 0) }}</td>
                    <td>{% if p.activo %}<span class="badge ok">Sí</span>{% else %}<span class="badge bad">No</span>{% endif %}</td>
                    <td>{{ p.updated_at }}</td>
                </tr>
                {% else %}
                <tr><td colspan="7" class="muted">Sin productos para mostrar.</td></tr>
                {% endfor %}
            </table>
        </div>
    </div>

    <div class="card">
        <h2>Historial de importaciones</h2>
        <div class="table-wrap">
            <table>
                <tr><th>Fecha</th><th>Archivo</th><th>Total</th><th>Creados</th><th>Actualizados</th><th>Errores</th><th>Usuario</th></tr>
                {% for h in history %}
                <tr>
                    <td>{{ h.created_at }}</td>
                    <td>{{ h.archivo_nombre }}</td>
                    <td>{{ h.total_filas }}</td>
                    <td>{{ h.creados }}</td>
                    <td>{{ h.actualizados }}</td>
                    <td>{{ h.errores }}</td>
                    <td>{{ h.usuario_nombre }}</td>
                </tr>
                {% else %}
                <tr><td colspan="7" class="muted">Sin historial.</td></tr>
                {% endfor %}
            </table>
        </div>
    </div>
    """, rows=rows, stats=stats, last_import=last_import, history=history, request=request)


@app.route("/productos/plantilla")
@login_required
@permission_required("productos")
def plantilla_productos():
    wb = Workbook()
    ws = wb.active
    ws.title = "Productos"
    headers = ["Código Producto", "Descripción", "Precio Compra Neto", "Precio Venta Bruto", "Stock", "Activo"]
    ws.append(headers)
    ws.append(["123456", "Ejemplo producto", 1000, 1990, 10, "SI"])
    for col in range(1, len(headers) + 1):
        ws.cell(row=1, column=col).font = ws.cell(row=1, column=col).font.copy(bold=True)
        ws.column_dimensions[ws.cell(row=1, column=col).column_letter].width = 24
    output = BytesIO()
    wb.save(output)
    output.seek(0)
    return send_file(
        output,
        as_attachment=True,
        download_name="plantilla_productos_cotizacion.xlsx",
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )


@app.route("/export/productos")
@login_required
@permission_required("productos")
def export_productos():
    rows = query_all("SELECT * FROM productos ORDER BY codigo_producto")
    columns = [
        ("codigo_producto", "Código Producto"),
        ("descripcion", "Descripción"),
        ("precio_compra_neto", "Precio Compra Neto"),
        ("precio_venta_bruto", "Precio Venta Bruto"),
        ("stock", "Stock"),
        ("activo", "Activo"),
        ("updated_at", "Actualizado"),
    ]
    return excel_response(f"productos_{today_str()}.xlsx", "Productos", columns, rows)


# ============================================================
# VENTAS / COTIZACIÓN IA ELIAS
# ============================================================


@app.route("/ventas", methods=["GET", "POST"])
@login_required
@permission_required("ventas")
def ventas():
    sesion = obtener_sesion_ventas()

    if request.method == "POST":
        action = request.form.get("action", "chat")
        cliente = request.form.get("cliente", "").strip()
        telefono = request.form.get("telefono", "").strip()

        if cliente or telefono:
            execute("UPDATE ventas_chat_sesiones SET cliente=?, telefono=?, updated_at=? WHERE id=?",
                    (cliente, telefono, now_str(), sesion["id"]))

        if action == "nueva_conversacion":
            session.pop("ventas_chat_sesion_id", None)
            flash("Nueva conversación iniciada con Elias.", "success")
            return redirect(url_for("ventas"))

        if action == "chat":
            mensaje = request.form.get("mensaje", "").strip()
            imagen = request.files.get("imagen_pedido")

            if imagen and imagen.filename and not imagen.mimetype.startswith("image/"):
                flash("El archivo adjunto debe ser una imagen.", "error")
                return redirect(url_for("ventas"))

            if not mensaje and not (imagen and imagen.filename):
                flash("Escribe un mensaje o adjunta una imagen.", "error")
                return redirect(url_for("ventas"))

            contenido_usuario = mensaje or "Imagen adjunta para analizar."
            if imagen and imagen.filename:
                contenido_usuario += f"\n[Imagen adjunta: {secure_filename(imagen.filename)}]"

            guardar_mensaje_ventas(sesion["id"], "user", contenido_usuario, "", 1 if imagen and imagen.filename else 0)
            contexto = contexto_conversacion_ventas(sesion["id"])
            respuesta, fuente = respuesta_chat_elias(mensaje, contexto, imagen)
            guardar_mensaje_ventas(sesion["id"], "assistant", respuesta, fuente, 0)

            flash("Elias respondió. Revisa el pedido antes de generar la cotización.", "success")
            return redirect(url_for("ventas"))

        if action == "generar_cotizacion":
            if query_one("SELECT COUNT(*) c FROM productos")["c"] == 0:
                flash("Primero debes importar la maestra de productos antes de cotizar.", "error")
                return redirect(url_for("ventas"))

            contexto = contexto_conversacion_ventas(sesion["id"])
            if not contexto.strip():
                flash("Primero conversa con Elias o pega la lista del cliente.", "error")
                return redirect(url_for("ventas"))

            data, fuente, raw = extraer_items_desde_sesion_ventas(sesion["id"])
            items = data.get("items", []) if isinstance(data, dict) else []
            if not items:
                flash("Elias no detectó productos cotizables en la conversación.", "error")
                return redirect(url_for("ventas"))

            cot_id = crear_cotizacion_desde_items(cliente, telefono, contexto, items, fuente, raw, sesion["id"])
            guardar_mensaje_ventas(
                sesion["id"],
                "quote",
                resumen_cotizacion_chat(cot_id),
                json.dumps({"cotizacion_id": cot_id}, ensure_ascii=False),
                0
            )
            flash("Cotización generada dentro del chat. Puedes seguir conversando o descargar PDF.", "success")
            return redirect(url_for("ventas") + "#chat-bottom")

    sesion = obtener_sesion_ventas()
    mensajes = mensajes_sesion_ventas_render(sesion["id"])
    last_import = query_one("SELECT * FROM producto_importaciones ORDER BY id DESC LIMIT 1")
    recent = query_all("SELECT * FROM cotizaciones ORDER BY id DESC LIMIT 15")
    stats = {
        "productos_activos": query_one("SELECT COUNT(*) c FROM productos WHERE activo = 1")["c"],
        "cotizaciones": query_one("SELECT COUNT(*) c FROM cotizaciones")["c"],
        "api_ok": openai_is_configured(),
        "modelo": openai_model_name(),
    }

    return render_page("Ventas", r"""
    <div class="page-head">
        <div>
            <h1>Ventas / Cotización IA</h1>
            <p>Chat de ventas con Elias. La conversación baja como chat normal y la cotización aparece dentro del mismo flujo.</p>
        </div>
    </div>

    <div class="grid grid-3">
        <div class="stat"><span>Productos activos</span><strong>{{ stats.productos_activos }}</strong></div>
        <div class="stat"><span>Cotizaciones generadas</span><strong>{{ stats.cotizaciones }}</strong></div>
        <div class="stat"><span>Modelo IA</span><strong style="font-size:18px;">{{ stats.modelo }}</strong></div>
    </div>

    <div class="card" style="margin-top:18px;">
        <div class="section-title">
            <div>
                <h2>Estado de maestra de productos</h2>
                {% if last_import %}
                <p class="muted">Última actualización: <b>{{ last_import.created_at }}</b> · {{ last_import.archivo_nombre }} · {{ last_import.total_filas }} filas.</p>
                {% else %}
                <p class="muted">No hay maestra de productos importada. Debe importarse antes de cotizar.</p>
                {% endif %}
            </div>
            {% if has_perm("productos") %}<a class="btn btn-secondary btn-small" href="{{ url_for('productos') }}">Actualizar productos</a>{% endif %}
        </div>
    </div>

    <div class="sales-layout" style="margin-top:18px;">
        <div class="ai-panel">
            <h2 style="margin-top:0;">Elias · Chat de ventas</h2>
            <p class="subtle">Pega la lista, conversa y corrige el pedido con Elias. Cuando esté listo, genera la cotización sin salir del chat.</p>

            {% if not stats.api_ok %}
                <div class="flash error">OPENAI_API_KEY no está configurada. Elias funcionará con extracción local básica.</div>
            {% endif %}

            <div class="chat-box" id="chatBox">
                {% if mensajes %}
                    {% for m in mensajes %}
                        {% if m.rol == 'quote' and m.quote %}
                            {% set q = m.quote['cot'] %}
                            <div class="chat-msg quote">
                                <div class="quote-chat-head">
                                    <div>
                                        <strong>Cotización {{ q.numero }}</strong><br>
                                        <small>{{ q.created_at }} · {{ q.estado }}</small>
                                    </div>
                                    <div style="text-align:right;">
                                        <strong>{{ q.subtotal_bruto|money }}</strong><br>
                                        <small>Total bruto confirmado</small>
                                    </div>
                                </div>
                                <div class="quote-chat-body">
                                    {% if m.quote['revisar_count'] > 0 %}
                                    <div class="flash error" style="margin-bottom:10px;">
                                        {{ m.quote['revisar_count'] }} línea(s) requieren revisión. No se sumaron al total si no tuvieron match confiable.
                                    </div>
                                    {% endif %}

                                    <div class="quote-chat-totals">
                                        <div class="quote-chip"><span>Total bruto</span><strong>{{ q.subtotal_bruto|money }}</strong></div>
                                        <div class="quote-chip"><span>Venta neta</span><strong>{{ q.venta_neta_total|money }}</strong></div>
                                        <div class="quote-chip"><span>Contribución</span><strong>{{ q.contribucion_total|money }}</strong></div>
                                        <div class="quote-chip"><span>Margen</span><strong>{{ q.margen_total_pct|percent }}</strong></div>
                                    </div>

                                    <table class="quote-table-mini">
                                        <tr><th>Estado</th><th>Código</th><th>Producto</th><th>Cant.</th><th>P. Unit.</th><th>Valor</th></tr>
                                        {% for i in m.quote['items'][:8] %}
                                        <tr>
                                            <td>
                                                {% if i.encontrado %}<span class="match-ok">OK</span>
                                                {% else %}<span class="match-review">REVISAR</span>{% endif %}
                                            </td>
                                            <td>{{ i.codigo_producto }}</td>
                                            <td>
                                                {% if i.encontrado %}
                                                    {{ i.descripcion_producto }}
                                                {% else %}
                                                    <b>{{ i.descripcion_solicitada }}</b><br><span class="muted">{{ i.descripcion_producto }}</span>
                                                {% endif %}
                                            </td>
                                            <td>{{ "%.2f"|format(i.cantidad or 0) }}</td>
                                            <td>{{ i.precio_venta_bruto|money }}</td>
                                            <td>{{ i.subtotal_bruto|money }}</td>
                                        </tr>
                                        {% endfor %}
                                    </table>

                                    {% if m.quote['items']|length > 8 %}
                                        <p class="muted" style="margin:8px 0 0;">Se muestran 8 de {{ m.quote['items']|length }} líneas. Abre el detalle para revisar todo.</p>
                                    {% endif %}

                                    <div class="actions" style="margin-top:12px;">
                                        <a class="btn btn-secondary btn-small" href="{{ url_for('ver_cotizacion', cotizacion_id=q.id) }}">Ver detalle</a>
                                        <a class="btn btn-primary btn-small" href="{{ url_for('pdf_cotizacion', cotizacion_id=q.id) }}">PDF</a>
                                        <a class="btn btn-secondary btn-small" href="{{ url_for('export_cotizacion', cotizacion_id=q.id) }}">Excel</a>
                                    </div>
                                </div>
                            </div>
                        {% else %}
                            <div class="chat-msg {{ 'user' if m.rol == 'user' else 'assistant' }}">{{ m.contenido }}</div>
                        {% endif %}
                    {% endfor %}
                {% else %}
                    <div class="chat-msg assistant">Hola, soy Elias. Pégame la lista del cliente o adjunta una imagen. Primero ordenamos el pedido y después puedes generar la cotización con el botón lateral.</div>
                {% endif %}
                <div id="chat-bottom"></div>
            </div>

            <form method="post" enctype="multipart/form-data" class="chat-compose">
                <input type="hidden" name="action" value="chat">
                <div>
                    <label>Mensaje para Elias</label>
                    <textarea name="mensaje" id="mensajeElias" placeholder="Pega aquí la lista del cliente o continúa la conversación..."></textarea>
                </div>
                <div style="margin-top:12px;">
                    <label>Imagen de lista o pedido</label>
                    <input type="file" name="imagen_pedido" accept="image/*">
                </div>
                <div class="actions">
                    <button class="btn" style="background:linear-gradient(135deg,#14b8a6,#2563eb);color:white;">Enviar a Elias</button>
                    <button class="btn btn-secondary" name="action" value="nueva_conversacion" formnovalidate>Nueva conversación</button>
                </div>
            </form>

            <script>
                (function(){
                    const box = document.getElementById("chatBox");
                    if (box) { box.scrollTop = box.scrollHeight; }
                    const txt = document.getElementById("mensajeElias");
                    if (txt) { txt.focus(); }
                })();
            </script>
        </div>

        <div class="side-card">
            <h3 style="margin-top:0;">Generar cotización</h3>
            <p class="muted">La cotización se mostrará dentro del chat. Luego puedes descargar PDF o Excel y seguir conversando con Elias.</p>
            <form method="post">
                <input type="hidden" name="action" value="generar_cotizacion">
                <div style="margin-bottom:12px;">
                    <label>Cliente</label>
                    <input name="cliente" value="{{ sesion.cliente or '' }}" placeholder="Opcional">
                </div>
                <div style="margin-bottom:12px;">
                    <label>Teléfono / WhatsApp</label>
                    <input name="telefono" value="{{ sesion.telefono or '' }}" placeholder="Opcional">
                </div>
                <button class="btn btn-primary" style="width:100%;justify-content:center;">Generar cotización en el chat</button>
            </form>

            <div class="placeholder" style="margin-top:16px;">
                <b>Regla de seguridad</b><br>
                Elias conversa y ordena. La cotización se crea solo con este botón. No se aceptan matches débiles como venta válida.
            </div>
        </div>
    </div>

    <div class="card" style="margin-top:18px;">
        <h2>Cotizaciones recientes</h2>
        <div class="table-wrap">
            <table>
                <tr><th>Número</th><th>Fecha</th><th>Cliente</th><th>Estado</th><th>Total bruto</th><th>Contribución</th><th>Margen</th><th>Usuario</th><th>Acción</th></tr>
                {% for c in recent %}
                <tr>
                    <td>{{ c.numero }}</td>
                    <td>{{ c.created_at }}</td>
                    <td>{{ c.cliente }}</td>
                    <td><span class="badge {% if c.estado == 'Requiere revisión' %}warn{% else %}ok{% endif %}">{{ c.estado }}</span></td>
                    <td>{{ c.subtotal_bruto|money }}</td>
                    <td>{{ c.contribucion_total|money }}</td>
                    <td>{{ c.margen_total_pct|percent }}</td>
                    <td>{{ c.usuario_nombre }}</td>
                    <td><a class="btn btn-secondary btn-small" href="{{ url_for('ver_cotizacion', cotizacion_id=c.id) }}">Ver</a></td>
                </tr>
                {% else %}
                <tr><td colspan="9" class="muted">Sin cotizaciones.</td></tr>
                {% endfor %}
            </table>
        </div>
    </div>
    """, last_import=last_import, recent=recent, stats=stats, mensajes=mensajes,
       sesion=sesion, has_perm=has_perm)


@app.route("/cotizaciones/<int:cotizacion_id>")
@login_required
@permission_required("ventas")
def ver_cotizacion(cotizacion_id):
    cot = query_one("SELECT * FROM cotizaciones WHERE id = ?", (cotizacion_id,))
    if not cot:
        flash("Cotización no encontrada.", "error")
        return redirect(url_for("ventas"))
    items = query_all("SELECT * FROM cotizacion_items WHERE cotizacion_id = ? ORDER BY id", (cotizacion_id,))
    revisar_count = query_one("SELECT COUNT(*) c FROM cotizacion_items WHERE cotizacion_id=? AND (encontrado=0 OR requiere_revision=1)", (cotizacion_id,))["c"]

    return render_page("Cotización", r"""
    <div class="page-head">
        <div>
            <h1>Cotización {{ cot.numero }}</h1>
            <p>Generada por {{ cot.usuario_nombre }} el {{ cot.created_at }} · Fuente: {{ cot.fuente }} · Estado: <b>{{ cot.estado }}</b></p>
        </div>
        <div class="actions">
            <a class="btn btn-secondary" href="{{ url_for('ventas') }}">Volver a ventas</a>
            <a class="btn btn-primary" href="{{ url_for('pdf_cotizacion', cotizacion_id=cot.id) }}">Descargar PDF</a>
            <a class="btn btn-secondary" href="{{ url_for('export_cotizacion', cotizacion_id=cot.id) }}">Exportar Excel</a>
        </div>
    </div>

    {% if revisar_count > 0 %}
    <div class="flash error">
        Hay {{ revisar_count }} línea(s) que requieren revisión. No se sumaron al total si no tuvieron match confiable.
    </div>
    {% endif %}

    <div class="grid grid-4">
        <div class="quote-total"><span class="muted">Total bruto confirmado</span><h2>{{ cot.subtotal_bruto|money }}</h2></div>
        <div class="quote-total"><span class="muted">Venta neta</span><h2>{{ cot.venta_neta_total|money }}</h2></div>
        <div class="quote-total"><span class="muted">Contribución</span><h2>{{ cot.contribucion_total|money }}</h2></div>
        <div class="quote-total"><span class="muted">Margen</span><h2>{{ cot.margen_total_pct|percent }}</h2></div>
    </div>

    <div class="card" style="margin-top:18px;">
        <h2>Detalle</h2>
        <div class="table-wrap">
            <table>
                <tr>
                    <th>Estado</th><th>Código</th><th>Solicitado</th><th>Producto encontrado / candidato</th>
                    <th>Cant.</th><th>Stock</th><th>Compra neto</th><th>Venta bruto</th>
                    <th>Subtotal bruto</th><th>Contribución</th><th>Margen</th><th>Score</th>
                </tr>
                {% for i in items %}
                <tr>
                    <td>
                        {% if i.encontrado %}
                            <span class="match-ok">OK</span>
                        {% elif i.requiere_revision %}
                            <span class="match-review">REVISAR</span>
                        {% else %}
                            <span class="match-bad">No encontrado</span>
                        {% endif %}
                    </td>
                    <td>{{ i.codigo_producto }}</td>
                    <td>{{ i.descripcion_solicitada }}</td>
                    <td>{{ i.descripcion_producto }}<br><span class="muted">{{ i.observacion }}</span></td>
                    <td>{{ "%.2f"|format(i.cantidad or 0) }}</td>
                    <td>{{ "%.2f"|format(i.stock or 0) }}</td>
                    <td>{{ i.precio_compra_neto|money }}</td>
                    <td>{{ i.precio_venta_bruto|money }}</td>
                    <td>{{ i.subtotal_bruto|money }}</td>
                    <td>{{ i.contribucion_total|money }}</td>
                    <td>{{ i.margen_pct|percent }}</td>
                    <td>
                        {% if i.match_score is not none %}
                            <span class="confidence {% if i.encontrado %}ok{% elif i.requiere_revision %}review{% else %}bad{% endif %}">{{ "%.1f"|format(i.match_score or 0) }}</span>
                        {% endif %}
                    </td>
                </tr>
                {% endfor %}
            </table>
        </div>
    </div>

    <div class="grid grid-2">
        <div class="card">
            <h3>Conversación / solicitud original</h3>
            <pre style="white-space:pre-wrap;font-family:inherit;">{{ cot.texto_original }}</pre>
        </div>
        <div class="card">
            <h3>Notas de control</h3>
            <p class="muted">Esta versión no cotiza automáticamente. Elias conversa primero; la cotización se genera solo al presionar el botón.</p>
            <p class="muted">Los productos con match débil quedan como REVISAR y no afectan el total bruto.</p>
            <p class="muted">IVA usado: {{ iva }}%</p>
        </div>
    </div>
    """, cot=cot, items=items, iva=round(iva_rate()*100, 2), revisar_count=revisar_count)


@app.route("/cotizaciones/<int:cotizacion_id>/pdf")
@login_required
@permission_required("ventas")
def pdf_cotizacion(cotizacion_id):
    cot = query_one("SELECT * FROM cotizaciones WHERE id = ?", (cotizacion_id,))
    if not cot:
        flash("Cotización no encontrada.", "error")
        return redirect(url_for("ventas"))

    items = query_all("""
        SELECT * FROM cotizacion_items
        WHERE cotizacion_id = ? AND encontrado = 1
        ORDER BY id
    """, (cotizacion_id,))
    revisar_count = query_one(
        "SELECT COUNT(*) c FROM cotizacion_items WHERE cotizacion_id=? AND (encontrado=0 OR requiere_revision=1)",
        (cotizacion_id,)
    )["c"]

    try:
        from xml.sax.saxutils import escape as xml_escape
        from reportlab.lib import colors
        from reportlab.lib.pagesizes import A4
        from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
        from reportlab.lib.units import cm
        from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
    except Exception as exc:
        flash(f"No se pudo generar PDF porque falta ReportLab: {exc}", "error")
        return redirect(url_for("ver_cotizacion", cotizacion_id=cotizacion_id))

    buffer = BytesIO()
    doc = SimpleDocTemplate(
        buffer,
        pagesize=A4,
        leftMargin=1.2*cm,
        rightMargin=1.2*cm,
        topMargin=1.2*cm,
        bottomMargin=1.2*cm,
    )
    styles = getSampleStyleSheet()
    title_style = ParagraphStyle(
        "TitleFerreteria",
        parent=styles["Title"],
        fontSize=18,
        leading=22,
        textColor=colors.HexColor("#0f172a"),
        alignment=0,
    )
    small_style = ParagraphStyle(
        "Small",
        parent=styles["Normal"],
        fontSize=8,
        leading=10,
        textColor=colors.HexColor("#334155"),
    )
    normal = styles["Normal"]

    story = []
    story.append(Paragraph("FERRETERÍA CLOUD TOOL", title_style))
    story.append(Paragraph("Cotización comercial generada por Elias / RUZ AI Systems", small_style))
    story.append(Spacer(1, 8))

    header_data = [
        ["N° Cotización", cot["numero"], "Fecha", cot["created_at"]],
        ["Cliente", cot["cliente"] or "Cliente", "Teléfono", cot["telefono"] or ""],
        ["Vendedor", cot["usuario_nombre"] or "", "Estado", cot["estado"] or ""],
    ]
    header_table = Table(header_data, colWidths=[3.1*cm, 6.1*cm, 2.4*cm, 6.0*cm])
    header_table.setStyle(TableStyle([
        ("GRID", (0,0), (-1,-1), 0.4, colors.HexColor("#cbd5e1")),
        ("BACKGROUND", (0,0), (0,-1), colors.HexColor("#f1f5f9")),
        ("BACKGROUND", (2,0), (2,-1), colors.HexColor("#f1f5f9")),
        ("FONTNAME", (0,0), (0,-1), "Helvetica-Bold"),
        ("FONTNAME", (2,0), (2,-1), "Helvetica-Bold"),
        ("FONTSIZE", (0,0), (-1,-1), 8),
        ("VALIGN", (0,0), (-1,-1), "TOP"),
        ("LEFTPADDING", (0,0), (-1,-1), 6),
        ("RIGHTPADDING", (0,0), (-1,-1), 6),
    ]))
    story.append(header_table)
    story.append(Spacer(1, 12))

    data = [["Código", "Descripción", "Cantidad", "Precio Unit.", "Valor"]]
    for i in items:
        data.append([
            i["codigo_producto"] or "",
            Paragraph(xml_escape(str(i["descripcion_producto"] or "")), small_style),
            f"{float(i['cantidad'] or 0):,.2f}".replace(",", "X").replace(".", ",").replace("X", "."),
            format_clp(i["precio_venta_bruto"]),
            format_clp(i["subtotal_bruto"]),
        ])

    if len(data) == 1:
        data.append(["", "Sin líneas confirmadas. Revisar cotización interna.", "", "", ""])

    table = Table(data, colWidths=[3.0*cm, 8.3*cm, 2.1*cm, 2.7*cm, 2.7*cm], repeatRows=1)
    table.setStyle(TableStyle([
        ("BACKGROUND", (0,0), (-1,0), colors.HexColor("#0f766e")),
        ("TEXTCOLOR", (0,0), (-1,0), colors.white),
        ("FONTNAME", (0,0), (-1,0), "Helvetica-Bold"),
        ("GRID", (0,0), (-1,-1), 0.35, colors.HexColor("#d1d5db")),
        ("FONTSIZE", (0,0), (-1,-1), 8),
        ("VALIGN", (0,0), (-1,-1), "TOP"),
        ("ALIGN", (2,1), (-1,-1), "RIGHT"),
        ("ROWBACKGROUNDS", (0,1), (-1,-1), [colors.white, colors.HexColor("#f8fafc")]),
    ]))
    story.append(table)
    story.append(Spacer(1, 12))

    totals_data = [
        ["Neto", format_clp(cot["venta_neta_total"])],
        [f"IVA {round(iva_rate()*100, 2)}%", format_clp(float(cot["subtotal_bruto"] or 0) - float(cot["venta_neta_total"] or 0))],
        ["Total", format_clp(cot["subtotal_bruto"])],
    ]
    totals = Table(totals_data, colWidths=[4*cm, 4*cm], hAlign="RIGHT")
    totals.setStyle(TableStyle([
        ("GRID", (0,0), (-1,-1), 0.4, colors.HexColor("#cbd5e1")),
        ("BACKGROUND", (0,0), (0,-1), colors.HexColor("#f1f5f9")),
        ("FONTNAME", (0,0), (-1,-1), "Helvetica-Bold"),
        ("FONTSIZE", (0,0), (-1,-1), 9),
        ("ALIGN", (1,0), (1,-1), "RIGHT"),
        ("BACKGROUND", (0,2), (-1,2), colors.HexColor("#dcfce7")),
    ]))
    story.append(totals)
    story.append(Spacer(1, 12))

    condiciones = (
        "CONDICIONES COMERCIALES<br/>"
        "1.- Validez de la oferta: 3 días.<br/>"
        "2.- Plazo de entrega: sujeto a disponibilidad de stock.<br/>"
        "3.- No incluye flete ni embalaje, salvo indicación expresa."
    )
    if revisar_count:
        condiciones += f"<br/><b>Nota interna:</b> {revisar_count} línea(s) quedaron pendientes de revisión y no fueron incluidas en este PDF comercial."
    story.append(Paragraph(condiciones, small_style))

    doc.build(story)
    buffer.seek(0)
    return send_file(
        buffer,
        as_attachment=True,
        download_name=f"{cot['numero']}.pdf",
        mimetype="application/pdf"
    )



@app.route("/cotizaciones/<int:cotizacion_id>/excel")
@login_required
@permission_required("ventas")
def export_cotizacion(cotizacion_id):
    cot = query_one("SELECT * FROM cotizaciones WHERE id = ?", (cotizacion_id,))
    if not cot:
        flash("Cotización no encontrada.", "error")
        return redirect(url_for("ventas"))

    items = query_all("SELECT * FROM cotizacion_items WHERE cotizacion_id = ? ORDER BY id", (cotizacion_id,))

    wb = Workbook()
    ws = wb.active
    ws.title = "Cotización"
    ws.append(["Cotización", cot["numero"]])
    ws.append(["Fecha", cot["created_at"]])
    ws.append(["Cliente", cot["cliente"] or ""])
    ws.append(["Teléfono", cot["telefono"] or ""])
    ws.append([])
    headers = [
        "Código", "Solicitado", "Producto", "Cantidad", "Stock", "Compra Neto",
        "Venta Bruto", "Subtotal Bruto", "Contribución", "Margen"
    ]
    ws.append(headers)

    for item in items:
        ws.append([
            item["codigo_producto"],
            item["descripcion_solicitada"],
            item["descripcion_producto"],
            item["cantidad"],
            item["stock"],
            item["precio_compra_neto"],
            item["precio_venta_bruto"],
            item["subtotal_bruto"],
            item["contribucion_total"],
            item["margen_pct"],
        ])

    ws.append([])
    ws.append(["", "", "", "", "", "", "Total bruto", cot["subtotal_bruto"], "Contribución", cot["contribucion_total"]])
    ws.append(["", "", "", "", "", "", "Venta neta", cot["venta_neta_total"], "Margen", cot["margen_total_pct"]])

    for row in ws.iter_rows(min_row=6, max_row=6):
        for cell in row:
            cell.font = cell.font.copy(bold=True)

    for col in ws.columns:
        letter = col[0].column_letter
        max_len = min(max(len(str(c.value or "")) for c in col) + 2, 42)
        ws.column_dimensions[letter].width = max_len

    output = BytesIO()
    wb.save(output)
    output.seek(0)

    return send_file(
        output,
        as_attachment=True,
        download_name=f"{cot['numero']}.xlsx",
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )


# ============================================================
# INTEGRACIONES / FACTURACIÓN.CL
# ============================================================

@app.route("/facturacion")
@login_required
@permission_required("integraciones")
def facturacion():
    return render_page("Facturación.cl", r"""
    <div class="page-head">
        <div>
            <h1>Integración Facturación.cl</h1>
            <p>Módulo preparado para conectar documentos electrónicos y evitar doble digitación.</p>
        </div>
    </div>

    <div class="card">
        <h2>Estado del módulo</h2>
        <div class="placeholder">
            Esta versión deja preparada la sección de integración, pero aún no consulta la API real.
            El siguiente paso técnico es conectar credenciales, WSDL/API y mapear documento → despacho.
        </div>
    </div>

    <div class="grid grid-2">
        <div class="card">
            <h3>Datos que debería traer</h3>
            <ul>
                <li>Factura, boleta o guía.</li>
                <li>PDF del documento.</li>
                <li>Monto total.</li>
                <li>Cliente y RUT.</li>
                <li>Fecha de emisión.</li>
                <li>Estado del documento.</li>
            </ul>
        </div>
        <div class="card">
            <h3>Uso operacional</h3>
            <ul>
                <li>Buscar documento por número.</li>
                <li>Precargar cliente y monto en despacho.</li>
                <li>Evitar documentos duplicados.</li>
                <li>Adjuntar link o PDF al registro.</li>
            </ul>
        </div>
    </div>
    """)


# ============================================================
# EXPORTACIONES EXCEL
# ============================================================

def excel_response(filename, sheet_title, columns, rows):
    wb = Workbook()
    ws = wb.active
    ws.title = sheet_title[:31]
    ws.append([label for key, label in columns])
    for row in rows:
        ws.append([row[key] if key in row.keys() else "" for key, label in columns])

    for col in ws.columns:
        max_len = 12
        col_letter = col[0].column_letter
        for cell in col:
            try:
                max_len = max(max_len, len(str(cell.value or "")) + 2)
            except Exception:
                pass
        ws.column_dimensions[col_letter].width = min(max_len, 42)

    output = BytesIO()
    wb.save(output)
    output.seek(0)
    return send_file(
        output,
        as_attachment=True,
        download_name=filename,
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )


@app.route("/exportar")
@login_required
@permission_required("exportar")
def exportar():
    return render_page("Exportar Excel", r"""
    <div class="page-head">
        <div>
            <h1>Exportar Excel / Reportes</h1>
            <p>Descarga datos operacionales y administrativos para respaldo o análisis.</p>
        </div>
    </div>

    <div class="grid grid-3">
        <a class="card" href="{{ url_for('export_despachos') }}"><h3>Despachos</h3><p class="muted">Registros de documentos, estados, clientes, usuarios y montos.</p></a>
        <a class="card" href="{{ url_for('export_mantenciones') }}"><h3>Mantenciones</h3><p class="muted">Historial de mantención, maquinaria, costo y responsable.</p></a>
        <a class="card" href="{{ url_for('export_auditoria') }}"><h3>Auditoría</h3><p class="muted">Eventos de edición y control por usuario.</p></a>
        <a class="card" href="{{ url_for('export_maquinarias') }}"><h3>Maquinarias</h3><p class="muted">Listado de maquinaria y estados.</p></a>
        <a class="card" href="{{ url_for('export_vehiculos') }}"><h3>Vehículos</h3><p class="muted">Patentes, documentos y vencimientos.</p></a>
        <a class="card" href="{{ url_for('export_productos') }}"><h3>Productos</h3><p class="muted">Maestra de precios, stock, margen y estado activo.</p></a>
        <a class="card" href="{{ url_for('ventas') }}"><h3>Cotizaciones</h3><p class="muted">Cotizaciones generadas por Elias y ventas.</p></a>
        <a class="card danger-zone" href="{{ url_for('backup_db') }}"><h3>Backup base de datos</h3><p class="muted">Descarga un respaldo completo de usuarios, despachos, maquinarias, mantenciones, productos, cotizaciones, auditoría y configuración.</p></a>
    </div>
    """)




@app.route("/backup-db")
@login_required
@permission_required("administracion")
def backup_db():
    backup_path = backup_database_if_exists("manual")
    if not backup_path:
        flash("No se pudo crear respaldo. Verifica que exista base de datos y Disk persistente.", "error")
        return redirect(url_for("exportar"))

    write_audit("backup", "descargar", None, "database", DB_PATH, backup_path)
    return send_file(
        backup_path,
        as_attachment=True,
        download_name=f"backup_ferreteria_cloud_tool_{today_str()}.db",
        mimetype="application/octet-stream"
    )


@app.route("/export/despachos")
@login_required
@permission_required("consulta")
def export_despachos():
    q = request.args.get("q", "").strip()
    estado = request.args.get("estado", "").strip()
    desde = request.args.get("desde", "").strip()
    hasta = request.args.get("hasta", "").strip()

    wheres = []
    params = []
    if q:
        wheres.append("(numero_documento LIKE ? OR tipo_documento LIKE ? OR usuario_nombre LIKE ?)")
        params += [f"%{q}%", f"%{q}%", f"%{q}%"]
    if estado:
        wheres.append("estado = ?")
        params.append(estado)
    if desde:
        wheres.append("date(created_at) >= date(?)")
        params.append(desde)
    if hasta:
        wheres.append("date(created_at) <= date(?)")
        params.append(hasta)

    where_sql = "WHERE " + " AND ".join(wheres) if wheres else ""
    rows = query_all(f"SELECT * FROM despachos {where_sql} ORDER BY id DESC", params)
    columns = [
        ("id", "ID"),
        ("numero_documento", "Número documento"),
        ("tipo_documento", "Tipo documento"),
        ("estado", "Estado"),
        ("monto", "Monto"),
        ("usuario_nombre", "Usuario"),
        ("created_at", "Creado"),
        ("updated_at", "Actualizado"),
    ]
    return excel_response(f"despachos_rapidos_{today_str()}.xlsx", "Despachos", columns, rows)

@app.route("/export/mantenciones")
@login_required
@permission_required("mantenciones")
def export_mantenciones():
    rows = query_all("SELECT * FROM mantenciones ORDER BY fecha DESC, id DESC")
    columns = [
        ("id", "ID"), ("maquinaria_nombre", "Maquinaria"), ("fecha", "Fecha"), ("tipo_mantencion", "Tipo"),
        ("estado", "Estado"), ("responsable", "Responsable"), ("costo", "Costo"),
        ("observaciones", "Observaciones"), ("usuario_nombre", "Usuario"), ("created_at", "Creado")
    ]
    return excel_response(f"mantenciones_{today_str()}.xlsx", "Mantenciones", columns, rows)


@app.route("/export/auditoria")
@login_required
@permission_required("auditoria")
def export_auditoria():
    rows = query_all("SELECT * FROM audit_log ORDER BY id DESC")
    columns = [
        ("id", "ID"), ("created_at", "Fecha"), ("modulo", "Módulo"), ("accion", "Acción"),
        ("registro_id", "Registro ID"), ("campo", "Campo"), ("valor_anterior", "Valor anterior"),
        ("valor_nuevo", "Valor nuevo"), ("usuario_nombre", "Usuario")
    ]
    return excel_response(f"auditoria_{today_str()}.xlsx", "Auditoria", columns, rows)


@app.route("/export/maquinarias")
@login_required
@permission_required("maquinarias")
def export_maquinarias():
    rows = query_all("SELECT * FROM maquinarias ORDER BY nombre")
    columns = [
        ("id", "ID"), ("codigo", "Código"), ("nombre", "Nombre"), ("tipo", "Tipo"), ("marca", "Marca"),
        ("modelo", "Modelo"), ("anio", "Año"), ("patente", "Patente"), ("estado", "Estado"),
        ("observaciones", "Observaciones"), ("created_at", "Creado")
    ]
    return excel_response(f"maquinarias_{today_str()}.xlsx", "Maquinarias", columns, rows)


@app.route("/export/vehiculos")
@login_required
@permission_required("vehiculos")
def export_vehiculos():
    rows = query_all("SELECT * FROM vehiculos ORDER BY patente")
    columns = [
        ("id", "ID"), ("patente", "Patente"), ("tipo", "Tipo"), ("marca", "Marca"), ("modelo", "Modelo"),
        ("anio", "Año"), ("estado", "Estado"), ("permiso_circulacion_vencimiento", "Permiso circulación"),
        ("revision_tecnica_vencimiento", "Revisión técnica"), ("seguro_obligatorio_vencimiento", "SOAP"),
        ("observaciones", "Observaciones"), ("created_at", "Creado")
    ]
    return excel_response(f"vehiculos_{today_str()}.xlsx", "Vehiculos", columns, rows)


# ============================================================
# ARRANQUE
# ============================================================

init_db()

if __name__ == "__main__":
    port = int(os.environ.get("PORT", 5000))
    app.run(host="0.0.0.0", port=port, debug=True)
